import os
from flask import Flask, flash, request, redirect, url_for
from werkzeug.utils import secure_filename

UPLOAD_FOLDER = '/Projects/Uploads'
ALLOWED_EXTENSIONS = {'gz'}

app = Flask(__name__)
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024 * 1024
# app.run(host='0.0.0.0', port=5000,debug=True)

def allowed_file(filename):
    return '.' in filename and \
           filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

@app.route('/', methods=['GET', 'POST'])
def upload_file():
    import datetime
    starttime = datetime.datetime.now()
    print(starttime)
    if request.method == 'POST':
        # check if the post request has the file part
        if 'file' not in request.files:
            flash('No file part')
            return redirect(request.url)
        file = request.files['file']
        # if user does not select file, browser also
        # submit an empty part without filename
        if file.filename == '':
            flash('No selected file')
            return redirect(request.url)
        if file and allowed_file(file.filename):
            filename = secure_filename(file.filename)
            file.save(os.path.join(app.config['UPLOAD_FOLDER'], filename))
            analyze_file(filename)
            excel(xlanalysisfile)
            endtime = datetime.datetime.now()
            print(endtime)
            return redirect(url_for('uploaded_file',filename=xlanalysisfilename))
    return '''
    <!doctype html>
    <head>
        <title>Upload gsXXXXX_snapshot.bak.gz File</title>
        <link href="static/css/style.css" rel="stylesheet" type="text/css" media="all">
        <link href="//fonts.googleapis.com/css?family=Oxygen:400,700" rel="stylesheet" type="text/css">
    </head>
    <body>
    <h1>OA Analysis</h1>
    <div class="login-form w3-agile">
        <h2>Upload a Snapshot File</h2>
        <form method=post enctype=multipart/form-data>
          <p>Select your gsXXXXX_snapshot.bak.gz</p>
          <input type=file name=file />
          <input type=submit value=Upload />
        </form>
    </div>
    </body>
    '''

def analyze_file(filename):
    import os
    from datetime import date
    import calendar
    import re
    import openpyxl
    import gzip,shutil
    month = calendar.month_abbr[date.today().month]
    year = date.today().year
    day = date.today().day
    username = os.environ["USERNAME"]
    #onedrivedir = "C:/Users/"+ username +"/OneDrive - National Information Solutions Cooperative/"
    downloadsdir = "C:/Users/"+ username +"/Downloads/"
    projectsdir = "C:/Projects/Uploads/"
    coopsdir = r"\\mofs\sdesvr\Support\MDM-DA\Coops"
    # companydir = str(number) + " " + name + "/"
    # monthdir = month + " " + str(year) + "/"
    global number
    number = re.search('([0-9])+', filename)
    number = number.group(0)

    #if not os.path.exists(onedrivedir + companydir):
    #    os.mkdir(onedrivedir + companydir)

    # if not os.path.exists(onedrivedir + companydir + monthdir):
        # os.mkdir(onedrivedir + companydir + monthdir)
        
    # if not os.path.exists(projectsdir + companydir):
        # os.mkdir(projectsdir + companydir)
        
    # if not os.path.exists(projectsdir + companydir + monthdir):
        # os.mkdir(projectsdir + companydir + monthdir)

    gzfilename, file_extension = os.path.splitext(filename)
    # if not os.path.exists(projectsdir + companydir + monthdir):
        # os.mkdir(projectsdir + companydir + monthdir)
        
    if file_extension == ".gz":
        with gzip.open(projectsdir + filename, 'r') as f_in, open(projectsdir + filename[:-3], 'wb') as f_out:
              shutil.copyfileobj(f_in, f_out)
    gzfile = projectsdir + gzfilename
    restore_db(gzfile, number, projectsdir)
    set_level(number)
    
    xltemplatefile = projectsdir + "OAMapWiseDataReview_Master_temp.xlsm"
    global xlanalysisfile
    xlanalysisfile = projectsdir + "gs" + str(number) + "-OAMapWiseDataReview-" + str(day) + str(month) + str(year) + ".xlsm"
    global xlanalysisfilename
    xlanalysisfilename = "gs" + str(number) + "-OAMapWiseDataReview-" + str(day) + str(month) + str(year) + ".xlsm"
    if os.path.exists(xltemplatefile):
        mywb = openpyxl.load_workbook(xltemplatefile, keep_vba=True)
        mywb.save(xlanalysisfile)
    else:
        print("File " + xltemplatefile + " not found.")
    
    
    return xlanalysisfilename, xlanalysisfile

def restore_db(gzfile, number, projectsdir):
    import pyodbc
    import os
    import time

    conn = pyodbc.connect("Driver={SQL Server};Server=.\SQLEXPRESS;Database=master;Trusted_Connection=yes", autocommit = True)
    conn.timeout = 60
    cursor = conn.cursor()
    file_list = []

    def get_filelistonly(bak_file):
        sqlcommand = r"""
                        RESTORE filelistonly FROM DISK = N'{bak_file}'
                     """.format(bak_file=bak_file)
        print(sqlcommand)
        cursor.execute(sqlcommand)
        rows = cursor.fetchall()

        for row in rows:
            fname = row[0]
            fext = os.path.splitext(row[1])[1]
            if "." not in fext:
                raise ValueError("No extension found in row")
            file_list.append({"fname": fname, "fext": fext})
        return file_list
    
    def get_drop_command(new_db):
        r = None
        if len(file_list) > 0:
            sqlcommand = r"""DROP DATABASE IF EXISTS {new_db}
                         """.format(new_db=new_db)
            r = sqlcommand
            try:
                cursor.execute(sqlcommand)
                while cursor.nextset():
                    pass
            except:
                print("Couldn't drop table")
                pass
        return r
    
    def get_restore_command(new_db, bak_file, file_list):
        r = None
        if len(file_list) > 0:
            sqlcommand = r"""RESTORE DATABASE {new_db} FROM DISK = N'{bak_file}'
                            WITH
                            FILE = 1,
                         """.format(new_db=new_db, bak_file=bak_file)
            sqlcommand = sqlcommand + ", \n".join(("MOVE N'{fname}' TO N'{projectsdir}\{new_db}{fext}'".format(fname=fl['fname'], fext=fl['fext'], new_db=new_db, number=number, projectsdir=projectsdir) for fl in file_list))
            sqlcommand = sqlcommand + ", NOUNLOAD, REPLACE, STATS = 5"
            r = sqlcommand
            sqlcommand = sqlcommand.replace("/" , "\\")
            try:
                cursor.execute(sqlcommand)
                while cursor.nextset():
                    pass
            except:
                print("Couldn't restore table")
                pass
        return r

    rows_empty = ()
    #backup_file = projectsdir + companydir + monthdir + gzfilename[:-3]
    file_list = get_filelistonly(gzfile)

    r = get_drop_command("gs" + number)
    time.sleep(15)
    r = get_restore_command("gs" + number, gzfile, file_list)
    time.sleep(15)
    conn.close()
    return

def set_level(number):
    import pyodbc

    conn = pyodbc.connect("Driver={SQL Server};Server=.\SQLEXPRESS;Database=master;Trusted_Connection=yes", autocommit = True)
    conn.timeout = 60
    cursor = conn.cursor()

    sqlcommand = r"""Use master
                    ALTER DATABASE gs{number}
                    SET COMPATIBILITY_LEVEL = 130;
                 """.format(number=number)
    cursor.execute(sqlcommand)
    while cursor.nextset():
        pass

    conn.close()
    return

def sumrows(table):
    try:
        cursor.execute("SELECT COUNT(*) FROM {table}".format(table=table))
        row = cursor.fetchone()
        totalrows = str(row[0])
        return totalrows
    except:
        return
    
def uniqueid(idcolumn, table):
    import pandas as pd
    c = "Error."
    facility_id = False
    for row in cursor.columns(table=table):
        if "gs_facility_id" in row:
            facility_id = True
    global cell_style 
    cell_style = 'Normal'
    global cell_alignment
    cell_alignment = 'False'
    rows = cursor.columns(table=table)
    if int(sumrows(table)) == 0:
                cell_style = 'Bad'
                c = "This table is empty."
    else:
        for row in cursor.columns(table=table):
            if idcolumn in row:
                r = r""
                sqlcommand = r.join(("SELECT {table}.OBJECTID, {table}.{idcolumn}".format(idcolumn=idcolumn,table=table)))
                if (facility_id == True and (not(idcolumn == "gs_facility_id"))):
                    sqlcommand = r.join((sqlcommand, """,{table}.gs_facility_id, Count({table}.OBJECTID) AS [Count] 
                        FROM {table} WHERE ({table}.{idcolumn} IS null OR {table}.{idcolumn} = '') 
                        GROUP BY {table}.{idcolumn}, {table}.OBJECTID""".format(idcolumn=idcolumn,table=table)))
                else:
                    sqlcommand = r.join((sqlcommand, """, Count({table}.OBJECTID) AS [Count] 
                        FROM {table} WHERE ({table}.{idcolumn} IS null OR {table}.{idcolumn} = '') 
                        GROUP BY {table}.{idcolumn}, {table}.OBJECTID""".format(idcolumn=idcolumn,table=table)))
                if (facility_id == True and (not(idcolumn == "gs_facility_id"))):
                    sqlcommand = r.join((sqlcommand, """,{table}.gs_facility_id 
                        ORDER BY {table}.{idcolumn}, Count({table}.OBJECTID) DESC, {table}.OBJECTID
                        """.format(idcolumn=idcolumn,table=table)))
                else:
                    sqlcommand = r.join((sqlcommand, """ 
                        ORDER BY {table}.{idcolumn}, Count({table}.OBJECTID) DESC, {table}.OBJECTID
                        """.format(idcolumn=idcolumn,table=table)))
                cursor.execute(sqlcommand)
                rows = cursor.fetchall()
                totalrows = sumrows(table)
                if not rows:
                    cell_style = 'Normal'
                    cell_alignment = 'False'
                    c = "All IDs are populated."
                else:
                    if (facility_id == True and (not(idcolumn == "gs_facility_id"))):
                        countid = str(sum(e[3] for e in rows))
                    else:
                        countid = str(sum(e[2] for e in rows))
                    percent = int(round(int(countid)/int(totalrows)*100))
                    cell_style = 'Bad'
                    c = str(countid) + " (" + str(percent) + "%)" + " blank values."
                    r = r""
                    sqlcommand = r.join(("SELECT {table}.OBJECTID, {table}.{idcolumn}".format(idcolumn=idcolumn,table=table)))
                    if (facility_id == True and (not(idcolumn == "gs_facility_id"))):
                        sqlcommand = r.join((sqlcommand, """,{table}.gs_facility_id, Count({table}.OBJECTID) AS [Count] 
                            FROM {table} WHERE ({table}.{idcolumn} IS null OR {table}.{idcolumn} = '') 
                            GROUP BY {table}.{idcolumn}, {table}.OBJECTID""".format(idcolumn=idcolumn,table=table)))
                    else:
                        sqlcommand = r.join((sqlcommand, """, Count({table}.OBJECTID) AS [Count] 
                            FROM {table} WHERE ({table}.{idcolumn} IS null OR {table}.{idcolumn} = '') 
                            GROUP BY {table}.{idcolumn}, {table}.OBJECTID""".format(idcolumn=idcolumn,table=table)))
                    if (facility_id == True and (not(idcolumn == "gs_facility_id"))):
                        sqlcommand = r.join((sqlcommand, """,{table}.gs_facility_id 
                            ORDER BY {table}.{idcolumn}, Count({table}.OBJECTID) DESC, {table}.OBJECTID
                            """.format(idcolumn=idcolumn,table=table)))
                    else:
                        sqlcommand = r.join((sqlcommand, """ 
                            ORDER BY {table}.{idcolumn}, Count({table}.OBJECTID) DESC, {table}.OBJECTID
                            """.format(idcolumn=idcolumn,table=table)))
                    df = pd.read_sql_query(sqlcommand, conn)
                    sheetname = category + "-Blanks"
                    createsheet(sheetname)
                    writedf(df, sheetname)
                    write_hyperlink(sheetname)
                break
            else:
                cell_style = 'Neutral'
                c = "This field needs to be added to the database."  
    return c
    
def duplicateid(idcolumn, idcolumn1, table):
    import pandas as pd
    c = "Error."
    global cell_style
    cell_style = 'Normal'
    global cell_alignment
    cell_alignment = 'False'
    rows = cursor.columns(table=table)
    for row in cursor.columns(table=table):
        if "gs_facility_id" in row:
            facility_id = True
        else:
            facility_id = False
    for row in cursor.columns(table=table):
        if "gs_bank_id" in row:
            bank_id = True
        else:
            bank_id = False
    if int(sumrows(table)) == 0:
        cell_style = 'Bad'
        c = "This table is empty."
    else:
        for row in cursor.columns(table=table):
            if idcolumn in row:
                for row in cursor.columns(table=table):
                    if idcolumn1 in row:
                        sqlcommand = r"""SELECT DISTINCT {table}.{idcolumn1}, {table}.{idcolumn}
                                        FROM {table}
                                        WHERE ((({table}.{idcolumn}) In (SELECT [{idcolumn}] FROM [{table}] As Tmp GROUP BY [{idcolumn}] HAVING Count(*)>1 )))
                                        ORDER BY {table}.{idcolumn};
                                    """.format(idcolumn=idcolumn,idcolumn1=idcolumn1,table=table)
                        cursor.execute(sqlcommand)
                        row = cursor.fetchall()
                        if not row:
                            cell_style = 'Normal'
                            cell_alignment = 'False'
                            c = "No duplicates found."
                        else:
                            cell_style = 'Bad'
                            c = str(len(row)) + " duplicates exist."
                            sqlcommand = r"""SELECT DISTINCT {table}.OBJECTID, {table}.{idcolumn}, {table}.{idcolumn1}
                                    FROM {table}
                                    WHERE ((({table}.{idcolumn}) In (SELECT [{idcolumn}] FROM [{table}] As Tmp GROUP BY [{idcolumn}] HAVING Count(*)>1 )))
                                    GROUP BY {table}.OBJECTID, {table}.{idcolumn}, {table}.{idcolumn1}
                                    ORDER BY {table}.{idcolumn}, {table}.OBJECTID, {table}.{idcolumn1}
                                    """.format(idcolumn=idcolumn,idcolumn1=idcolumn1,table=table)
                            r = r""
                            sqlcommand = r.join(("SELECT DISTINCT {table}.OBJECTID, {table}.{idcolumn}, {table}.{idcolumn1}".format(idcolumn=idcolumn,idcolumn1=idcolumn1,table=table)))
                            if (facility_id == True and (not(idcolumn == "gs_facility_id")) and bank_id == True):
                                sqlcommand = r.join((sqlcommand, """, {table}.gs_facility_id 
                                    FROM {table} 
                                    WHERE ((({table}.{idcolumn}) 
                                        In (SELECT [{idcolumn}] 
                                        FROM [{table}] As Tmp 
                                        GROUP BY [{idcolumn}] HAVING Count(*)>1 )))
                                    GROUP BY {table}.OBJECTID, {table}.{idcolumn}, {table}.{idcolumn1}, {table}.gs_facility_id, {table}.gs_bank_id
                                    ORDER BY {table}.{idcolumn}, {table}.gs_bank_id, {table}.{idcolumn1}, {table}.gs_facility_id, {table}.OBJECTID  """.format(idcolumn=idcolumn,idcolumn1=idcolumn1,table=table)))
                            else:
                                sqlcommand = r.join((sqlcommand, """ FROM {table} 
                                    WHERE ((({table}.{idcolumn}) 
                                        In (SELECT [{idcolumn}] 
                                        FROM [{table}] As Tmp 
                                        GROUP BY [{idcolumn}] HAVING Count(*)>1 )))
                                    GROUP BY {table}.OBJECTID, {table}.{idcolumn}, {table}.{idcolumn1}
                                    ORDER BY {table}.OBJECTID, {table}.{idcolumn}, {table}.{idcolumn1}""".format(idcolumn=idcolumn,idcolumn1=idcolumn1,table=table)))
                            df = pd.read_sql_query(sqlcommand, conn)
                            sheetname = category + "-" + idcolumn[len("gs_"):]
                            createsheet(sheetname)
                            writedf(df, sheetname)
                            write_hyperlink(sheetname)
                        break
                    else:
                        cell_style = 'Neutral'
                        c = "The field ''" + str(idcolumn1) + "' needs to be added to the database."
            else:
                cell_style = 'Neutral'
                c = "The field ''" + str(idcolumn) + "' needs to be added to the database."
    return c
    
def duplicate_xfmr(idcolumn, idcolumn1, table):
    import pandas as pd
    c = "Error."
    facility_id = False
    global cell_style
    cell_style = 'Normal'
    global cell_alignment
    cell_alignment = 'False'
    rows = cursor.columns(table=table)
    for row in cursor.columns(table=table):
        if "gs_facility_id" in row:
            facility_id = True
        else:
            facility_id = False
    if int(sumrows(table)) == 0:
        cell_style = 'Bad'
        c = "This table is empty."
    else:
        for row in cursor.columns(table=table):
            if idcolumn in row:
                for row in cursor.columns(table=table):
                    if idcolumn1 in row:
                        r = r""
                        sqlcommand = r.join(("SELECT {table}.OBJECTID, ".format(table=table)))
                        if (facility_id == True and (not(idcolumn == "gs_facility_id" or idcolumn1 == "gs_facility_id"))):
                            sqlcommand = r.join((sqlcommand, "{table}.gs_facility_id, {table}.{idcolumn}, {table}.{idcolumn1}".format(idcolumn=idcolumn,idcolumn1=idcolumn1,table=table)))
                        else:
                            sqlcommand = r.join((sqlcommand, "{table}.{idcolumn}, {table}.{idcolumn1}".format(idcolumn=idcolumn,idcolumn1=idcolumn1,table=table)))
                        if not ((idcolumn == "gs_phase") or (idcolumn1 == "gs_phase")):
                            sqlcommand = r.join((sqlcommand, ",{table}.gs_phase FROM {table} WHERE ((({table}.{idcolumn}) In (SELECT {idcolumn} FROM {table} As Tmp GROUP BY {idcolumn}, {idcolumn1}, gs_phase HAVING Count(*)>1 ))) GROUP BY {table}.OBJECTID, {table}.{idcolumn}, {table}.{idcolumn1}".format(idcolumn=idcolumn, idcolumn1=idcolumn1, table=table)))
                            if (facility_id == True and (not(idcolumn == "gs_facility_id" or idcolumn1 == "gs_facility_id"))):
                                sqlcommand = r.join((sqlcommand, ", {table}.gs_facility_id".format(table=table)))
                        else:
                            sqlcommand = r.join((sqlcommand, " FROM {table} WHERE ((({table}.{idcolumn}) In (SELECT {idcolumn} FROM {table} As Tmp GROUP BY {idcolumn}, {idcolumn1} HAVING Count(*)>1 ))) GROUP BY {table}.OBJECTID, {table}.{idcolumn}, {table}.{idcolumn1}".format(idcolumn=idcolumn, idcolumn1=idcolumn1, table=table)))
                            if (facility_id == True and (not(idcolumn == "gs_facility_id" or idcolumn1 == "gs_facility_id"))):
                                sqlcommand = r.join((sqlcommand, ", {table}.gs_facility_id".format(table=table)))
                        if not ((idcolumn == "gs_phase") or (idcolumn1 == "gs_phase")):
                            sqlcommand = r.join((sqlcommand, ", {table}.gs_phase ORDER BY {table}.{idcolumn}, {table}.OBJECTID, {table}.{idcolumn1}, {table}.gs_phase".format(idcolumn=idcolumn,idcolumn1=idcolumn1,table=table)))
                            if (facility_id == True and (not(idcolumn == "gs_facility_id" or idcolumn1 == "gs_facility_id"))):
                                sqlcommand = r.join((sqlcommand, ", {table}.gs_facility_id".format(table=table)))
                        else:
                            sqlcommand = r.join((sqlcommand," ORDER BY {table}.{idcolumn}, {table}.OBJECTID, {table}.{idcolumn1}".format(idcolumn=idcolumn,idcolumn1=idcolumn1,table=table)))
                            if (facility_id == True and (not(idcolumn == "gs_facility_id" or idcolumn1 == "gs_facility_id"))):
                                sqlcommand = r.join((sqlcommand, ", {table}.gs_facility_id".format(table=table)))
                        cursor.execute(sqlcommand)
                        row = cursor.fetchall()
                        if not row:
                            cell_style = 'Normal'
                            cell_alignment = 'False'
                            c = "No duplicates found."
                        else:
                            cell_style = 'Bad'
                            c = str(len(row)) + " duplicates exist."
                            r = r""
                            sqlcommand = r.join(("SELECT {table}.OBJECTID, ".format(table=table)))
                            if (facility_id == True and (not(idcolumn == "gs_facility_id" or idcolumn1 == "gs_facility_id"))):
                                sqlcommand = r.join((sqlcommand, "{table}.gs_facility_id, {table}.{idcolumn}, {table}.{idcolumn1}".format(idcolumn=idcolumn,idcolumn1=idcolumn1,table=table)))
                            else:
                                sqlcommand = r.join((sqlcommand, "{table}.{idcolumn}, {table}.{idcolumn1}".format(idcolumn=idcolumn,idcolumn1=idcolumn1,table=table)))
                            if not ((idcolumn == "gs_phase") or (idcolumn1 == "gs_phase")):
                                sqlcommand = r.join((sqlcommand, ",{table}.gs_phase FROM {table} WHERE ((({table}.{idcolumn}) In (SELECT {idcolumn} FROM {table} As Tmp GROUP BY {idcolumn}, {idcolumn1}, gs_phase HAVING Count(*)>1 ))) GROUP BY {table}.OBJECTID, {table}.{idcolumn}, {table}.{idcolumn1}".format(idcolumn=idcolumn, idcolumn1=idcolumn1, table=table)))
                                if (facility_id == True and (not(idcolumn == "gs_facility_id" or idcolumn1 == "gs_facility_id"))):
                                    sqlcommand = r.join((sqlcommand, ", {table}.gs_facility_id".format(table=table)))
                            else:
                                sqlcommand = r.join((sqlcommand, " FROM {table} WHERE ((({table}.{idcolumn}) In (SELECT {idcolumn} FROM {table} As Tmp GROUP BY {idcolumn}, {idcolumn1} HAVING Count(*)>1 ))) GROUP BY {table}.OBJECTID, {table}.{idcolumn}, {table}.{idcolumn1}".format(idcolumn=idcolumn, idcolumn1=idcolumn1, table=table)))
                                if (facility_id == True and (not(idcolumn == "gs_facility_id" or idcolumn1 == "gs_facility_id"))):
                                    sqlcommand = r.join((sqlcommand, ", {table}.gs_facility_id".format(table=table)))
                            if not ((idcolumn == "gs_phase") or (idcolumn1 == "gs_phase")):
                                sqlcommand = r.join((sqlcommand, ", {table}.gs_phase ORDER BY {table}.{idcolumn}, {table}.OBJECTID, {table}.{idcolumn1}, {table}.gs_phase".format(idcolumn=idcolumn,idcolumn1=idcolumn1,table=table)))
                                if (facility_id == True and (not(idcolumn == "gs_facility_id" or idcolumn1 == "gs_facility_id"))):
                                    sqlcommand = r.join((sqlcommand, ", {table}.gs_facility_id".format(table=table)))
                            else:
                                sqlcommand = r.join((sqlcommand," ORDER BY {table}.{idcolumn}, {table}.OBJECTID, {table}.{idcolumn1}".format(idcolumn=idcolumn,idcolumn1=idcolumn1,table=table)))
                                if (facility_id == True and (not(idcolumn == "gs_facility_id" or idcolumn1 == "gs_facility_id"))):
                                    sqlcommand = r.join((sqlcommand, ", {table}.gs_facility_id".format(table=table)))
                            df = pd.read_sql_query(sqlcommand, conn)
                            sheetname = category + "-" + idcolumn[len("gs_"):]
                            createsheet(sheetname)
                            writedf(df, sheetname)
                            write_hyperlink(sheetname)
                        break
                    else:
                        cell_style = 'Neutral'
                        c = "The field ''" + str(idcolumn1) + "' needs to be added to the database."
            else:
                cell_style = 'Neutral'
                c = "The field ''" + str(idcolumn) + "' needs to be added to the database."
    return c

def duplicate_system():
    import pandas as pd
    c = "Error."
    global cell_style
    cell_style = 'Normal'
    global cell_alignment
    cell_alignment = 'False'
    table = "gs_service_point"
    rows = cursor.columns(table=table)
    if int(sumrows(table)) == 0:
        cell_style = 'Bad'
        c = "This table is empty."
    else:
        r = r""
        sqlcommand = r.join((r"""SELECT t.id AS 'ID',COUNT(T.ID) AS 'Count'
        FROM (SELECT gs_service_map_location AS 'ID' FROM GS_SERVICE_POINT"""))
        table = "gs_transformer"
        for row in cursor.columns(table=table):
            if "gs_equipment_location" in row:
                sqlcommand = r.join((sqlcommand, """
                UNION ALL
                SELECT gs_equipment_location AS 'ID' FROM gs_transformer
                WHERE (((gs_transformer.gs_equipment_location) 
                    In (SELECT gs_equipment_location 
                    FROM gs_transformer As Tmp 
                    GROUP BY gs_equipment_location, gs_bank_id, gs_phase HAVING Count(*)>1 ))) """))
        table = "gs_overcurrent_device"
        for row in cursor.columns(table=table):
            if "gs_equipment_location" in row:
                sqlcommand = r.join((sqlcommand, """
                UNION ALL
                SELECT GS_EQUIPMENT_LOCATION AS 'ID' FROM GS_OVERCURRENT_DEVICE"""))
        table = "gs_switch"
        for row in cursor.columns(table=table):
            if "gs_equipment_location" in row:
                sqlcommand = r.join((sqlcommand, """
                UNION ALL
                SELECT GS_EQUIPMENT_LOCATION AS 'ID' FROM GS_SWITCH"""))
        table = "gs_capacitor_bank"
        for row in cursor.columns(table=table):
            if "gs_equipment_location" in row:
                sqlcommand = r.join((sqlcommand, """
                UNION ALL
                SELECT GS_EQUIPMENT_LOCATION AS 'ID' FROM GS_CAPACITOR_BANK"""))
        table = "gs_generator"
        for row in cursor.columns(table=table):
            if "gs_equipment_location" in row:
                sqlcommand = r.join((sqlcommand, """
                UNION ALL
                SELECT GS_EQUIPMENT_LOCATION AS 'ID' FROM GS_GENERATOR"""))
        table = "gs_motor"
        for row in cursor.columns(table=table):
            if "gs_equipment_location" in row:
                sqlcommand = r.join((sqlcommand, """
                UNION ALL
                SELECT GS_EQUIPMENT_LOCATION AS 'ID' FROM GS_MOTOR"""))
        # table = "gs_street_light"
        # for row in cursor.columns(table=table):
            # if "gs_equipment_location" in row:
                # sqlcommand = r.join((sqlcommand, """
                # UNION ALL
                # SELECT GS_EQUIPMENT_LOCATION AS 'ID' FROM GS_STREET_LIGHT"""))
        table = "gs_voltage_regulator"
        for row in cursor.columns(table=table):
            if "gs_equipment_location" in row:
                sqlcommand = r.join((sqlcommand, """
                UNION ALL
                SELECT GS_EQUIPMENT_LOCATION AS 'ID' FROM GS_VOLTAGE_REGULATOR"""))
        sqlcommand = r.join((sqlcommand, """) AS T group by t.id having count(*)>1"""))
        cursor.execute(sqlcommand)
        row = cursor.fetchall()
        if not row:
            cell_style = 'Normal'
            cell_alignment = 'False'
            c = "No duplicates found."
        else:
            cell_style = 'Bad'
            c = str(len(row)) + " duplicates exist."
            df = pd.read_sql_query(sqlcommand, conn)
            sheetname = category + "-SystemDuplicates"
            createsheet(sheetname)
            writedf(df, sheetname)
            write_hyperlink(sheetname)
    return c

def xfmr_voltage(idcolumn, idcolumn1, table):
    import pandas as pd
    c = "Error."
    facility_id = False
    global cell_style
    cell_style = 'Normal'
    global cell_alignment
    cell_alignment = 'False'
    rows = cursor.columns(table=table)
    for row in rows:
        if idcolumn in row:
            for row in cursor.columns(table=table):
                if idcolumn1 in row:
                    r = r""
                    sqlcommand = r"""SELECT {table}.{idcolumn}, {table}.{idcolumn1}, Count({table}.OBJECTID) AS [Count]
                    FROM {table}
                    GROUP BY {table}.{idcolumn}, {table}.{idcolumn1}
                    ORDER BY Count({table}.OBJECTID) DESC
                    """.format(idcolumn=idcolumn,idcolumn1=idcolumn1,table=table)
                    cursor.execute(sqlcommand)
                    row = cursor.fetchall()
                    df = pd.read_sql_query(sqlcommand, conn)
                    sheetname = category + "-VoltageTable"
                    createsheet(sheetname)
                    writedf(df, sheetname)
                    break
    return
    
def span_assembly(table):
    import pandas as pd
    c = "Error."
    facility_id = False
    global cell_style
    cell_style = 'Normal'
    global cell_alignment
    cell_alignment = 'False'
    rows = cursor.columns(table=table)
    for row in rows:
        if idcolumn in row:
            for row in cursor.columns(table=table):
                if idcolumn1 in row:
                    r = r""
                    sqlcommand = r"""SELECT gs_assembly_ref.gs_assembly_name, gs_assembly_ref.gs_engineering_analysis_name, Count({table}.gs_guid) AS [Count], {table}.gs_subtype_cd, gs_assembly_ref.gs_assembly_description
                    FROM ({table} INNER JOIN gs_attached_assemblies ON {table}.gs_guid = gs_attached_assemblies.gs_network_feature_guid) INNER JOIN gs_assembly_ref ON gs_attached_assemblies.gs_assembly_guid = gs_assembly_ref.gs_guid
                    GROUP BY gs_assembly_ref.gs_assembly_name, gs_assembly_ref.gs_engineering_analysis_name, {table}.gs_subtype_cd, gs_assembly_ref.gs_assembly_description, gs_assembly_ref.gs_is_flow_assembly
                    HAVING (((gs_assembly_ref.gs_is_flow_assembly)='true' Or (gs_assembly_ref.gs_is_flow_assembly)='true'))
                    ORDER BY gs_assembly_ref.gs_assembly_name, gs_assembly_ref.gs_engineering_analysis_name, Count({table}.gs_guid) DESC
                    """.format(idcolumn=idcolumn,idcolumn1=idcolumn1,table=table)
                    cursor.execute(sqlcommand)
                    row = cursor.fetchall()
                    df = pd.read_sql_query(sqlcommand, conn)
                    sheetname = category + "-AssemblyTable"
                    createsheet(sheetname)
                    writedf(df, sheetname)
                    df = df[df.duplicated(subset=['gs_assembly_name'], keep=False)]
                    sheetname = category + "-DupAssemblyTable"
                    createsheet(sheetname)
                    writedf(df, sheetname)
                    write_hyperlink(sheetname)
                    break
    return
    
def fieldsummary(idcolumn, table):
    import pandas as pd
    c = "Error."
    global cell_style 
    cell_style = 'Normal'
    global cell_alignment
    cell_alignment = 'False'
    facility_id = False
    for row in cursor.columns(table=table):
        if "gs_facility_id" in row:
            facility_id = True
    for row in cursor.columns(table=table):
        if "gs_equipment_location" in row:
            equipment_location = True
    for row in cursor.columns(table=table):
        if idcolumn in row:
            file_list = []
            if idcolumn == "gs_phase":
                sqlcommand = r"""SELECT {table}.{idcolumn}, Count({table}.OBJECTID) AS [Count]
                            FROM {table}
                            GROUP BY {table}.{idcolumn}
                            ORDER BY CASE WHEN {idcolumn} = 'NULL' THEN '1'
                                          WHEN {idcolumn} = 'UNK' THEN '2'
                                          ELSE {idcolumn} END ASC
                            """.format(idcolumn=idcolumn,table=table)
            elif idcolumn == "gs_installation_date":
                sqlcommand = r"""SELECT {table}.{idcolumn}, Count({table}.OBJECTID) AS [Count]
                            FROM {table}
                            GROUP BY {table}.{idcolumn}
                            """.format(idcolumn=idcolumn,table=table)
            else:
                sqlcommand = r"""SELECT {table}.{idcolumn}, Count({table}.OBJECTID) AS [Count]
                            FROM {table}
                            GROUP BY {table}.{idcolumn}
                            ORDER BY CASE WHEN {idcolumn} = TRY_CONVERT(numeric, 'NULL') THEN '1'
                                          WHEN {idcolumn} = TRY_CONVERT(numeric, 'UNK') THEN '2'
                                          ELSE {idcolumn} END ASC
                            """.format(idcolumn=idcolumn,table=table)
            print(idcolumn + " " + sqlcommand)
            cursor.execute(sqlcommand)
            row = cursor.fetchall()
            totalrows = sumrows(table)
            if not row:
                cell_style = 'Bad'
                c = "This table is empty."
            else:
                if int(sumrows(table)) == int(row[0][1]):
                    cell_style = 'Normal'
                    cell_alignment = 'False'
                    if row[0][0] == None:
                        c = "All fields are NULL."
                    elif row[0][0] == 0:
                        c = "All fields are populated with '0'."
                    else:
                        c = "All fields are populated with '" + str(row[0][0]) +"'."
                    if (row[0][0] == None) or (row[0][0] == "") or (row[0][0] == "UNK")or (row[0] == "0E-8"):
                            cell_style = 'Bad'
                            r = r""
                            sqlcommand = r.join(("SELECT {table}.OBJECTID, {table}.{idcolumn}".format(idcolumn=idcolumn,table=table)))
                            sqlcommand = r.join((sqlcommand, """ FROM {table}
                                 WHERE ({table}.{idcolumn} IS null 
                                 OR {table}.{idcolumn} = TRY_CONVERT(numeric, '') 
                                 OR {table}.{idcolumn} LIKE '%UNK%')
                                 GROUP BY {table}.OBJECTID, {table}.{idcolumn}
                                 ORDER BY {table}.OBJECTID, {table}.{idcolumn}""".format(idcolumn=idcolumn,table=table)))
                            df = pd.read_sql_query(sqlcommand, conn)
                            sheetname = category + "-" + idcolumn[len("gs_"):]
                            createsheet(sheetname)
                            writedf(df, sheetname)
                            write_hyperlink(sheetname)
                else:
                    cell_style = 'Normal'
                    for row in row:
                        if (row[0] == None) or (row[0] == "") or (row[0] == "UNK")or (row[0] == "0E-8"):
                            cell_style = 'Bad'
                        fcol = row[0]
                        if fcol == 0:
                            fcol = "0"
                        if fcol == None:
                            fcol = "NULL"
                        fnum = row[1]
                        fper = int(round(int(fnum)/int(totalrows)*100))
                        file_list.append({"fcol": fcol, "fnum": fnum, "fper": fper})
                    cell_alignment = 'True'
                    c = r""        
                    c = c.join(("{fnum} ({fper}%) populated with '{fcol}'.\n".format(fcol=fl['fcol'], fnum=fl['fnum'], fper=fl['fper']) for fl in file_list))
                    if cell_style == 'Bad':

                        r = r""
                        sqlcommand = r.join(("SELECT {table}.OBJECTID, {table}.{idcolumn}".format(idcolumn=idcolumn,table=table)))
                        if (facility_id == True and (not(idcolumn == "gs_facility_id")) and equipment_location == True and (not(idcolumn == "gs_equipment_location"))):
                            sqlcommand = r.join((sqlcommand, """,{table}.gs_equipment_location, {table}.gs_facility_id 
                                FROM {table}
                                WHERE ({table}.{idcolumn} IS null 
                                    OR {table}.{idcolumn} = TRY_CONVERT(numeric, '') 
                                    OR {table}.{idcolumn} LIKE '%UNK%')
                                    GROUP BY {table}.OBJECTID, {table}.{idcolumn},{table}.gs_equipment_location, {table}.gs_facility_id
                                    ORDER BY {table}.OBJECTID, {table}.{idcolumn},{table}.gs_equipment_location, {table}.gs_facility_id""".format(idcolumn=idcolumn,table=table)))
                        elif (facility_id == True and (not(idcolumn == "gs_facility_id"))):
                            sqlcommand = r.join((sqlcommand, """,{table}.gs_facility_id 
                                FROM {table}
                                WHERE ({table}.{idcolumn} IS null 
                                    OR {table}.{idcolumn} = TRY_CONVERT(numeric, '') 
                                    OR {table}.{idcolumn} LIKE '%UNK%')
                                    GROUP BY {table}.OBJECTID, {table}.{idcolumn}, {table}.gs_facility_id
                                    ORDER BY {table}.OBJECTID, {table}.{idcolumn}, {table}.gs_facility_id""".format(idcolumn=idcolumn,table=table))) 
                        else:
                            sqlcommand = r.join((sqlcommand, """ FROM {table}
                                WHERE ({table}.{idcolumn} IS null 
                                    OR {table}.{idcolumn} = TRY_CONVERT(numeric, '') 
                                    OR {table}.{idcolumn} LIKE '%UNK%')
                                    GROUP BY {table}.OBJECTID, {table}.{idcolumn}
                                    ORDER BY {table}.OBJECTID, {table}.{idcolumn}""".format(idcolumn=idcolumn,table=table)))
                        df = pd.read_sql_query(sqlcommand, conn)
                        sheetname = category + "-" + idcolumn[len("gs_"):]
                        createsheet(sheetname)
                        writedf(df, sheetname)
                        write_hyperlink(sheetname)
            break
        else:
            cell_style = 'Neutral'
            c = "This field needs to be added to the database."
    
    return c

def fieldsummarygt(idcolumn, idcolumn1, value, table):
    import pandas as pd
    c = "Error."
    global cell_style 
    cell_style = 'Normal'
    global cell_alignment
    cell_alignment = 'False'
    facility_id = False
    idcolumnFound = False
    idcolumn1Found = False						 						  
    for row in cursor.columns(table=table):
        if "gs_facility_id" in row:
            facility_id = True
    for row in cursor.columns(table=table):
        if idcolumn in row:
            idcolumnFound = True
    for row in cursor.columns(table=table):
        if idcolumn1 in row:
            idcolumn1Found = True
    for row in cursor.columns(table=table):
        if idcolumnFound == True and idcolumn1Found == True:						   
            file_list = []
            sqlcommand = r"""SELECT {table}.{idcolumn}, Count({table}.OBJECTID) AS [Count]
                        FROM {table}
                        WHERE TRY_CONVERT(numeric, {idcolumn1}) > {value}
                        GROUP BY {table}.{idcolumn}
                        ORDER BY CASE WHEN {idcolumn} = TRY_CONVERT(numeric, 'NULL') THEN '1'
                                      WHEN {idcolumn} = TRY_CONVERT(numeric, 'UNK') THEN '2'
                                      ELSE {idcolumn} END ASC
                        """.format(idcolumn=idcolumn,idcolumn1=idcolumn1,value=value,table=table)
            cursor.execute(sqlcommand)
            row = cursor.fetchall()
            totalrows = sumrows(table)
            if not row:
                cell_style = 'Bad'
                c = ""
            else:
                if int(sumrows(table)) == int(row[0][1]):
                    cell_style = 'Normal'
                    cell_alignment = 'False'
                    if row[0][0] == None:
                        c = "All fields are NULL."
                    elif row[0][0] == 0:
                        c = "All fields are populated with '0'."
                    else:
                        c = "All fields are populated with '" + str(row[0][0]) +"'."
                    if (row[0][0] == None) or (row[0][0] == "") or (row[0][0] == "UNK") or (row[0] == "0E-8"):
                            cell_style = 'Bad'
                else:
                    cell_style = 'Normal'
                    for row in row:
                        if (row[0] == None) or (row[0] == "") or (row[0] == "UNK") or (row[0] == "0E-8"):
                            cell_style = 'Bad'
                        fcol = row[0]
                        if fcol == 0:
                            fcol = "0"
                        if fcol == None:
                            fcol = "NULL"
                        fnum = row[1]
                        fper = int(round(int(fnum)/int(totalrows)*100))
                        file_list.append({"fcol": fcol, "fnum": fnum, "fper": fper})
                    cell_alignment = 'True'
                    c = r""        
                    c = c.join(("{fnum} ({fper}%) populated with '{fcol}'.\n".format(fcol=fl['fcol'], fnum=fl['fnum'], fper=fl['fper']) for fl in file_list))
                    if cell_style == 'Bad':

                        r = r""
                        sqlcommand = r.join(("SELECT {table}.OBJECTID, {table}.{idcolumn}".format(idcolumn=idcolumn,table=table)))
                        if (facility_id == True and (not(idcolumn == "gs_facility_id"))):
                            sqlcommand = r.join((sqlcommand, """,{table}.gs_facility_id 
                                FROM {table}
                                WHERE ({table}.{idcolumn} IS null 
                                    OR {table}.{idcolumn} = TRY_CONVERT(numeric, '') 
                                    OR {table}.{idcolumn} LIKE '%UNK%')
                                    GROUP BY {table}.OBJECTID, {table}.{idcolumn}, {table}.gs_facility_id
                                    ORDER BY {table}.OBJECTID, {table}.{idcolumn}, {table}.gs_facility_id""".format(idcolumn=idcolumn,table=table)))
                        else:
                            sqlcommand = r.join((sqlcommand, """ FROM {table}
                                WHERE ({table}.{idcolumn} IS null 
                                    OR {table}.{idcolumn} = TRY_CONVERT(numeric, '') 
                                    OR {table}.{idcolumn} LIKE '%UNK%')
                                    GROUP BY {table}.OBJECTID, {table}.{idcolumn}
                                    ORDER BY {table}.OBJECTID, {table}.{idcolumn}""".format(idcolumn=idcolumn,table=table)))
                        df = pd.read_sql_query(sqlcommand, conn)
                        sheetname = category + "-" + idcolumn[len("gs_"):]
                        createsheet(sheetname)
                        writedf(df, sheetname)
                        write_hyperlink(sheetname)
            break
        else:
            cell_style = 'Neutral'
            if idcolumnFound == False:
                c = idcolumn + " needs to be added to the database."
            elif idcolumn1Found == False:
                c = idcolumn1 + " needs to be added to the database."
    
    return c

def fieldsummaryeq(idcolumn, idcolumn1, value, table):
    import pandas as pd
    c = "Error."
    global cell_style 
    cell_style = 'Normal'
    global cell_alignment
    cell_alignment = 'False'
    facility_id = False
    idcolumnFound = False
    idcolumn1Found = False				 
    for row in cursor.columns(table=table):
        if "gs_facility_id" in row:
            facility_id = True
    for row in cursor.columns(table=table):
        if idcolumn in row:
            idcolumnFound = True
    for row in cursor.columns(table=table):
        if idcolumn1 in row:
            idcolumn1Found = True
    for row in cursor.columns(table=table):
        if idcolumnFound == True and idcolumn1Found == True:			
            file_list = []
            sqlcommand = r"""SELECT {table}.{idcolumn}, Count({table}.OBJECTID) AS [Count]
                        FROM {table}
                        WHERE TRY_CONVERT(numeric, {idcolumn1}) = {value}
                        GROUP BY {table}.{idcolumn}
                        ORDER BY CASE WHEN {idcolumn} = TRY_CONVERT(numeric, 'NULL') THEN '1'
                                      WHEN {idcolumn} = TRY_CONVERT(numeric, 'UNK') THEN '2'
                                      ELSE {idcolumn} END ASC
                        """.format(idcolumn=idcolumn,idcolumn1=idcolumn1,table=table,value=value)
            cursor.execute(sqlcommand)
            row = cursor.fetchall()
            totalrows = sumrows(table)
            if not row:
                cell_style = 'Bad'
                c = ""
            else:
                if int(sumrows(table)) == int(row[0][1]):
                    cell_style = 'Normal'
                    cell_alignment = 'False'
                    if row[0][0] == None:
                        c = "All fields are NULL."
                    elif row[0][0] == 0:
                        c = "All fields are populated with '0'."
                    else:
                        c = "All fields are populated with '" + str(row[0][0]) +"'."
                    if (row[0][0] == None) or (row[0][0] == "") or (row[0][0] == "UNK") or (row[0] == "0E-8"):
                            cell_style = 'Bad'
                else:
                    cell_style = 'Normal'
                    for row in row:
                        if (row[0] == None) or (row[0] == "") or (row[0] == "UNK") or (row[0] == "0E-8"):
                            cell_style = 'Bad'
                        fcol = row[0]
                        if fcol == 0:
                            fcol = "0"
                        if fcol == None:
                            fcol = "NULL"
                        fnum = row[1]
                        fper = int(round(int(fnum)/int(totalrows)*100))
                        file_list.append({"fcol": fcol, "fnum": fnum, "fper": fper})
                    cell_alignment = 'True'
                    c = r""        
                    c = c.join(("{fnum} ({fper}%) populated with '{fcol}'.\n".format(fcol=fl['fcol'], fnum=fl['fnum'], fper=fl['fper']) for fl in file_list))
                    if cell_style == 'Bad':

                        r = r""
                        sqlcommand = r.join(("SELECT {table}.OBJECTID, {table}.{idcolumn}".format(idcolumn=idcolumn,table=table)))
                        if (facility_id == True and (not(idcolumn == "gs_facility_id"))):
                            sqlcommand = r.join((sqlcommand, """,{table}.gs_facility_id 
                                FROM {table}
                                WHERE ({table}.{idcolumn} IS null 
                                    OR {table}.{idcolumn} = TRY_CONVERT(numeric, '') 
                                    OR {table}.{idcolumn} LIKE '%UNK%')
                                    GROUP BY {table}.OBJECTID, {table}.{idcolumn}, {table}.gs_facility_id
                                    ORDER BY {table}.OBJECTID, {table}.{idcolumn}, {table}.gs_facility_id""".format(idcolumn=idcolumn,table=table)))
                        else:
                            sqlcommand = r.join((sqlcommand, """ FROM {table}
                                WHERE ({table}.{idcolumn} IS null 
                                    OR {table}.{idcolumn} = TRY_CONVERT(numeric, '') 
                                    OR {table}.{idcolumn} LIKE '%UNK%')
                                    GROUP BY {table}.OBJECTID, {table}.{idcolumn}
                                    ORDER BY {table}.OBJECTID, {table}.{idcolumn}""".format(idcolumn=idcolumn,table=table)))
                        df = pd.read_sql_query(sqlcommand, conn)
                        sheetname = category + "-" + idcolumn[len("gs_"):]
                        createsheet(sheetname)
                        writedf(df, sheetname)
                        write_hyperlink(sheetname)
            break
        else:
            cell_style = 'Neutral'
            if idcolumn in row:
                c = idcolumn + " needs to be added to the database."
            elif idcolumn1 in row:
                c = idcolumn1 + " needs to be added to the database."
    
    return c

def fieldsummaryeqtext(idcolumn, idcolumn1, value, table):
    import pandas as pd
    c = "Error."
    global cell_style 
    cell_style = 'Normal'
    global cell_alignment
    cell_alignment = 'False'
    facility_id = False
    idcolumnFound = False
    idcolumn1Found = False
    for row in cursor.columns(table=table):
        if "gs_facility_id" in row:
            facility_id = True
    for row in cursor.columns(table=table):
        if idcolumn in row:
            idcolumnFound = True
    for row in cursor.columns(table=table):
        if idcolumn1 in row:
            idcolumn1Found = True
    for row in cursor.columns(table=table):
        if idcolumnFound == True and idcolumn1Found == True:
            file_list = []
            sqlcommand = r"""SELECT {table}.{idcolumn}, Count({table}.OBJECTID) AS [Count]
                        FROM {table}
                        WHERE {idcolumn1} = {value}
                        GROUP BY {table}.{idcolumn}
                        ORDER BY CASE WHEN {idcolumn} = TRY_CONVERT(numeric, 'NULL') THEN '1'
                                      WHEN {idcolumn} = TRY_CONVERT(numeric, 'UNK') THEN '2'
                                      ELSE {idcolumn} END ASC
                        """.format(idcolumn=idcolumn,idcolumn1=idcolumn1,table=table,value=value)
            cursor.execute(sqlcommand)
            row = cursor.fetchall()
            totalrows = sumrows(table)
            if not row:
                c = ""
            else:
                if int(sumrows(table)) == int(row[0][1]):
                    cell_style = 'Normal'
                    cell_alignment = 'False'
                    if row[0][0] == None:
                        c = "All fields are NULL."
                    elif row[0][0] == 0:
                        c = "All fields are populated with '0'."
                    else:
                        c = "All fields are populated with '" + str(row[0][0]) +"'."
                    if (row[0][0] == None) or (row[0][0] == "") or (row[0][0] == "UNK") or (row[0] == "0E-8"):
                            cell_style = 'Bad'
                else:
                    cell_style = 'Normal'
                    for row in row:
                        if (row[0] == None) or (row[0] == "") or (row[0] == "UNK") or (row[0] == "0E-8"):
                            cell_style = 'Bad'
                        fcol = row[0]
                        if fcol == 0:
                            fcol = "0"
                        if fcol == None:
                            fcol = "NULL"
                        fnum = row[1]
                        fper = int(round(int(fnum)/int(totalrows)*100))
                        file_list.append({"fcol": fcol, "fnum": fnum, "fper": fper})
                    cell_alignment = 'True'
                    c = r""        
                    c = c.join(("{fnum} ({fper}%) populated with '{fcol}'.\n".format(fcol=fl['fcol'], fnum=fl['fnum'], fper=fl['fper']) for fl in file_list))
                    if cell_style == 'Bad':

                        r = r""
                        sqlcommand = r.join(("SELECT {table}.OBJECTID, {table}.{idcolumn}".format(idcolumn=idcolumn,table=table)))
                        if (facility_id == True and (not(idcolumn == "gs_facility_id"))):
                            sqlcommand = r.join((sqlcommand, """,{table}.gs_facility_id 
                                FROM {table}
                                WHERE ({table}.{idcolumn} IS null 
                                    OR {table}.{idcolumn} = TRY_CONVERT(numeric, '') 
                                    OR {table}.{idcolumn} LIKE '%UNK%')
                                    GROUP BY {table}.OBJECTID, {table}.{idcolumn}, {table}.gs_facility_id
                                    ORDER BY {table}.OBJECTID, {table}.{idcolumn}, {table}.gs_facility_id""".format(idcolumn=idcolumn,table=table)))
                        else:
                            sqlcommand = r.join((sqlcommand, """ FROM {table}
                                WHERE ({table}.{idcolumn} IS null 
                                    OR {table}.{idcolumn} = TRY_CONVERT(numeric, '') 
                                    OR {table}.{idcolumn} LIKE '%UNK%')
                                    GROUP BY {table}.OBJECTID, {table}.{idcolumn}
                                    ORDER BY {table}.OBJECTID, {table}.{idcolumn}""".format(idcolumn=idcolumn,table=table)))
                        df = pd.read_sql_query(sqlcommand, conn)
                        sheetname = category + "-" + idcolumn[len("gs_"):]
                        createsheet(sheetname)
                        writedf(df, sheetname)
                        write_hyperlink(sheetname)
            break
        else:
            cell_style = 'Neutral'
            if idcolumn in row:
                c = idcolumn + " needs to be added to the database."
            elif idcolumn1 in row:
                c = idcolumn1 + " needs to be added to the database."
    
    return c

def missingfield(idcolumn, table):
    c = "Error."
    global cell_style 
    cell_style = 'Normal'
    global cell_alignment
    cell_alignment = 'False'
    for row in cursor.columns(table=table):
        if idcolumn in row:
            file_list = []

            sqlcommand = r"""SELECT {table}.{idcolumn}, Count({table}.OBJECTID) AS [Count]
                        FROM {table}
                        WHERE ({table}.{idcolumn} IS null OR {table}.{idcolumn} = '')
                        GROUP BY {table}.{idcolumn}
                        ORDER BY {table}.{idcolumn}, Count({table}.OBJECTID) DESC
                        """.format(idcolumn=idcolumn,table=table)
            cursor.execute(sqlcommand)
            row = cursor.fetchall()
            totalrows = sumrows(table)
            
            try:
                if int(sumrows(table)) == int(row[0][1]):
                    cell_style = 'Normal'
                    cell_alignment = 'False'
                    if row[0][0] == None:
                        c = "All fields are NULL."
                    elif row[0][0] == 0:
                        c = "All fields are populated with '0'."
                    else:
                        c = "All fields are populated with '" + str(row[0][0]) +"'."
                    if (row[0][0] == None) or (row[0][0] == "") or (row[0][0] == "UNK"):
                                cell_style = 'Bad'
                else:
                    for row in row:
                        if (row[0] == None) or (row[0] == "") or (row[0] == "UNK"):
                            cell_style = 'Bad'
                        fcol = row[0]
                        if fcol == 0:
                            fcol = "0"
                        if fcol == None:
                            fcol = "NULL"
                        fnum = row[1]
                        fper = int(round(int(fnum)/int(totalrows)*100))
                        file_list.append({"fcol": fcol, "fnum": fnum, "fper": fper})
                    cell_alignment = 'True'
                    c = r""        
                    c = c.join(("{fnum} ({fper}%) populated with '{fcol}'.\n".format(fcol=fl['fcol'], fnum=fl['fnum'], fper=fl['fper']) for fl in file_list))
                break
            except IndexError:
                break
        else:
            cell_style = 'Neutral'
            c = "This field needs to be added to the database."
    
    return c
    
def neutral(idcolumn, table):
    import pandas as pd
    c = "Error."
    global cell_style 
    cell_style = 'Normal'
    global cell_alignment
    cell_alignment = 'False'
    for row in cursor.columns(table=table):
        if idcolumn in row:
            file_list = []

            sqlcommand = r"""SELECT {table}.{idcolumn}, {table}.gs_subtype_cd, Count({table}.OBJECTID) AS [Count]
                        FROM {table}
                        WHERE {table}.gs_subtype_cd =  1
                        AND ({table}.{idcolumn} IS null OR {table}.{idcolumn} = '' OR {table}.{idcolumn} LIKE '%UNK%')
                        GROUP BY {table}.{idcolumn}, {table}.gs_subtype_cd
                        ORDER BY {table}.{idcolumn}, {table}.gs_subtype_cd, Count({table}.OBJECTID) DESC
                        """.format(idcolumn=idcolumn,table=table)
            cursor.execute(sqlcommand)
            row = cursor.fetchall()
            totalrows = sumrows(table)
            
            try:
                if int(sumrows(table)) == int(row[0][2]):
                    cell_style = 'Normal'
                    cell_alignment = 'False'
                    if row[0][0] == None:
                        c = "All fields are NULL."
                    if row[0][0] == 0:
                        c = "All fields are populated with '0'."
                    else:
                        c = "All fields are populated with '" + str(row[0][0]) +"'."
                    if (row[0][0] == None) or (row[0][0] == "") or (row[0][0] == "UNK"):
                        cell_style = 'Bad'
                        sqlcommand = r"""SELECT {table}.OBJECTID, {table}.{idcolumn}
                                    FROM {table}
                                    WHERE {table}.gs_subtype_cd =  1
                                    AND ({table}.{idcolumn} IS null OR {table}.{idcolumn} = '' OR {table}.{idcolumn} LIKE '%UNK%')
                                    GROUP BY {table}.OBJECTID, {table}.{idcolumn}
                                    ORDER BY {table}.OBJECTID, {table}.{idcolumn}
                                    """.format(idcolumn=idcolumn,table=table)
                        df = pd.read_sql_query(sqlcommand, conn)
                        sheetname = category + "-" + idcolumn[len("gs_"):]
                        createsheet(sheetname)
                        writedf(df, sheetname)

                else:
                    for row in row:
                        if (row[0] == None) or (row[0] == "") or (row[0] == "UNK"):
                            cell_style = 'Bad'
                        fcol = row[0]
                        if fcol == 0:
                            fcol = "0"
                        if fcol == None:
                            fcol = "NULL"
                        fnum = row[2]
                        fper = int(round(int(fnum)/int(totalrows)*100))
                        file_list.append({"fcol": fcol, "fnum": fnum, "fper": fper})
                    cell_alignment = 'True'
                    c = r""        
                    c = c.join(("{fnum} ({fper}%) populated with '{fcol}'.\n".format(fcol=fl['fcol'], fnum=fl['fnum'], fper=fl['fper']) for fl in file_list))
                    if cell_style == 'Bad':
                        sqlcommand = r"""SELECT {table}.OBJECTID, {table}.{idcolumn}
                                    FROM {table}
                                    WHERE {table}.gs_subtype_cd =  1
                                    AND ({table}.{idcolumn} IS null OR {table}.{idcolumn} ='' OR {table}.{idcolumn} LIKE '%UNK%')
                                    GROUP BY {table}.OBJECTID, {table}.{idcolumn}
                                    ORDER BY {table}.OBJECTID, {table}.{idcolumn}
                                    """.format(idcolumn=idcolumn,table=table)
                        df = pd.read_sql_query(sqlcommand, conn)
                        sheetname = category + "-" + idcolumn[len("gs_"):]
                        createsheet(sheetname)
                        writedf(df, sheetname)
                        write_hyperlink(sheetname)
                break
            except IndexError:
                break
        else:
            cell_style = 'Neutral'
            c = "This field needs to be added to the database."
    
    return c
    
def nullabc(idcolumn, idcolumn1, idcolumn2, idcolumn3, table):
    import pandas as pd
    global cell_style 
    cell_style = 'Normal'
    global cell_alignment
    cell_alignment = 'False'
    facility_id = False
    for row in cursor.columns(table=table):
        if "gs_facility_id" in row:
            facility_id = True
    for row in cursor.columns(table=table):
        if "gs_equipment_location" in row:
            equipment_loc = True
    totalrows = sumrows(table)
    sqlcommand = r"""SELECT {table}.{idcolumn1}, {table}.{idcolumn}
                    FROM {table}
                    WHERE {table}.{idcolumn} LIKE '%a%'
                    AND ({table}.{idcolumn1} IS null OR {table}.{idcolumn1} = TRY_CONVERT(varchar,'') OR {table}.{idcolumn1} LIKE TRY_CONVERT(varchar,'%fake%') OR {table}.{idcolumn1} LIKE TRY_CONVERT(varchar,'%unk%'))
                    ORDER BY {table}.{idcolumn};
                """.format(idcolumn=idcolumn,idcolumn1=idcolumn1,table=table)

    cursor.execute(sqlcommand)
    row = cursor.fetchall()
    anum = len(row)
    aper = round(anum/int(totalrows)*100)
    c = str(anum) + " (" + str(aper) + "%) 'A', "

    sqlcommand = r"""SELECT {table}.{idcolumn2}, {table}.{idcolumn}
                    FROM {table}
                    WHERE {table}.{idcolumn} LIKE '%b%'
                    AND ({table}.{idcolumn2} IS null OR {table}.{idcolumn2} = TRY_CONVERT(varchar,'') OR {table}.{idcolumn2} LIKE TRY_CONVERT(varchar,'%fake%') OR {table}.{idcolumn2} LIKE TRY_CONVERT(varchar,'%unk%'))
                    ORDER BY {table}.{idcolumn};
                """.format(idcolumn=idcolumn,idcolumn2=idcolumn2,table=table)

    
    cursor.execute(sqlcommand)
    row = cursor.fetchall()
    bnum = len(row)
    bper = round(bnum/int(totalrows)*100)
    c = c + str(bnum) + " (" + str(bper) + "%) 'B', "
    
    sqlcommand = r"""SELECT {table}.{idcolumn3}, {table}.{idcolumn}
                    FROM {table}
                    WHERE {table}.{idcolumn} LIKE '%c%'
                    AND ({table}.{idcolumn3} IS null OR {table}.{idcolumn3} = TRY_CONVERT(varchar,'') OR {table}.{idcolumn3} LIKE TRY_CONVERT(varchar,'%fake%') OR {table}.{idcolumn3} LIKE TRY_CONVERT(varchar,'%unk%'))
                    ORDER BY {table}.{idcolumn};
                """.format(idcolumn=idcolumn,idcolumn3=idcolumn3,table=table)

    
    cursor.execute(sqlcommand)
    row = cursor.fetchall()
    cnum = len(row)
    cper = round(cnum/int(totalrows)*100)
    c = c + str(cnum) + " (" + str(cper) + "%) 'C' are not populated. "
    
    sqlcommand = r"""SELECT {table}.{idcolumn3}, {table}.{idcolumn}
                FROM {table}
                WHERE ({table}.{idcolumn} LIKE '%a%' OR {table}.{idcolumn} LIKE '%b%' OR {table}.{idcolumn} LIKE '%c%')
                AND ({table}.{idcolumn1} IS null OR {table}.{idcolumn1} = TRY_CONVERT(varchar,'') OR {table}.{idcolumn1} LIKE TRY_CONVERT(varchar,'%fake%') OR {table}.{idcolumn1} LIKE TRY_CONVERT(varchar,'%unk%'))
                AND ({table}.{idcolumn2} IS null OR {table}.{idcolumn2} = TRY_CONVERT(varchar,'') OR {table}.{idcolumn2} LIKE TRY_CONVERT(varchar,'%fake%') OR {table}.{idcolumn2} LIKE TRY_CONVERT(varchar,'%unk%'))
                AND ({table}.{idcolumn3} IS null OR {table}.{idcolumn3} = TRY_CONVERT(varchar,'') OR {table}.{idcolumn3} LIKE TRY_CONVERT(varchar,'%fake%') OR {table}.{idcolumn3} LIKE TRY_CONVERT(varchar,'%unk%'))
                ORDER BY {table}.{idcolumn};
            """.format(idcolumn=idcolumn,idcolumn1=idcolumn1,idcolumn2=idcolumn2,idcolumn3=idcolumn3,table=table)

    cursor.execute(sqlcommand)
    row = cursor.fetchall()
    tnum = len(row)
    tper = round(tnum/int(totalrows)*100)
    c = c + str(tnum) + " (" + str(tper) + "%) total."
    
    if anum + bnum + cnum == 0:
        c = "All are populated."
    else:
        cell_style = 'Bad'
        r = r""
        sqlcommand = r.join(("SELECT {table}.OBJECTID,{table}.{idcolumn},{table}.{idcolumn1},{table}.{idcolumn2},{table}.{idcolumn3}".format(idcolumn=idcolumn,idcolumn1=idcolumn1,idcolumn2=idcolumn2,idcolumn3=idcolumn3,table=table)))
        if (facility_id == True):
            sqlcommand = r.join((sqlcommand, ",{table}.gs_equipment_location".format(table=table)))
        if (facility_id == True):
            sqlcommand = r.join((sqlcommand, ",{table}.gs_facility_id".format(table=table)))
        sqlcommand = r.join((sqlcommand, """ FROM {table}
                WHERE ({table}.{idcolumn} LIKE '%a%' OR {table}.{idcolumn} LIKE '%b%' OR {table}.{idcolumn} LIKE '%c%')
                AND ({table}.{idcolumn1} IS null OR {table}.{idcolumn1} = TRY_CONVERT(varchar,'') OR {table}.{idcolumn1} LIKE TRY_CONVERT(varchar,'%fake%') OR {table}.{idcolumn1} LIKE TRY_CONVERT(varchar,'%unk%'))
                AND ({table}.{idcolumn2} IS null OR {table}.{idcolumn2} = TRY_CONVERT(varchar,'') OR {table}.{idcolumn2} LIKE TRY_CONVERT(varchar,'%fake%') OR {table}.{idcolumn2} LIKE TRY_CONVERT(varchar,'%unk%'))
                AND ({table}.{idcolumn3} IS null OR {table}.{idcolumn3} = TRY_CONVERT(varchar,'') OR {table}.{idcolumn3} LIKE TRY_CONVERT(varchar,'%fake%') OR {table}.{idcolumn3} LIKE TRY_CONVERT(varchar,'%unk%'))
                ORDER BY {table}.{idcolumn};
                """.format(idcolumn=idcolumn,idcolumn1=idcolumn1,idcolumn2=idcolumn2,idcolumn3=idcolumn3,table=table)))
        df = pd.read_sql_query(sqlcommand, conn)
        sheetname = category + "-" + idcolumn1[len("gs_"):] + ",b,c"
        createsheet(sheetname)
        writedf(df, sheetname)
        write_hyperlink(sheetname)

    return c
    
def createsheet(sheetname):
    if sheetname in wb.sheetnames:
        #wb.remove(wb[sheetname])
        #wb.create_sheet(sheetname)
        return
    else:
        wb.create_sheet(sheetname)
    return
    
def writedf(df, sheetname):
    from openpyxl.utils.dataframe import dataframe_to_rows
    ws = wb[sheetname]
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)
    for cell in ws[1]:
        cell.style = '40 % - Accent3'
    #ws.column_dimensions['A'].auto_size = True
    #ws.column_dimensions['B'].auto_size = True
    #ws.column_dimensions['C'].auto_size = True
    #ws.column_dimensions['D'].auto_size = True
    createsheet("Tasklist")
    ws = wb["Tasklist"]
    ws.append(tuple([sheetname]))
    ws = wb["Overview"]
    return
    
def write_hyperlink(sheetname):
    hyperlink = "#'" + sheetname + "'!A1"
    ws['D'+ str(cell)].hyperlink = hyperlink
    ws['D'+ str(cell)].value = sheetname
    ws['D'+ str(cell)].style = "Hyperlink"
    return

def is_feeder_bay(idcolumn, table):
    import pandas as pd
    c = "Error."
    facility_id = False
    for row in cursor.columns(table=table):
        if "gs_facility_id" in row:
            facility_id = True
    global cell_style 
    cell_style = 'Normal'
    global cell_alignment
    cell_alignment = 'False'
    rows = cursor.columns(table=table)
    if int(sumrows(table)) == 0:
                cell_style = 'Bad'
                c = "This table is empty."
    else:
        for row in cursor.columns(table=table):
            if idcolumn in row:
                r = r""
                sqlcommand = r.join(("SELECT {table}.OBJECTID, {table}.{idcolumn}".format(idcolumn=idcolumn,table=table)))
                if (facility_id == True and (not(idcolumn == "gs_facility_id"))):
                    sqlcommand = r.join((sqlcommand, """,{table}.gs_facility_id, {table}.gs_overcurrent_device_subtype, {table}.gs_is_feeder_bay
                        FROM {table} 
                        WHERE ({table}.{idcolumn} = '4' AND ({table}.gs_is_feeder_bay <> '1' OR {table}.gs_is_feeder_bay IS NULL)) OR ({table}.gs_overcurrent_device_subtype <> '4' AND {table}.gs_is_feeder_bay <> '0') 
                        GROUP BY {table}.{idcolumn}, {table}.OBJECTID, {table}.gs_facility_id, {table}.gs_overcurrent_device_subtype, {table}.gs_is_feeder_bay""".format(idcolumn=idcolumn,table=table)))
                else:
                    sqlcommand = r.join((sqlcommand, """, {table}.gs_overcurrent_device_subtype, {table}.gs_is_feeder_bay
                        FROM {table} 
                        WHERE ({table}.{idcolumn} = '4' AND ({table}.gs_is_feeder_bay <> '1' OR {table}.gs_is_feeder_bay IS NULL)) OR ({table}.gs_overcurrent_device_subtype <> '4' AND {table}.gs_is_feeder_bay <> '0') 
                        GROUP BY {table}.{idcolumn}, {table}.OBJECTID, {table}.gs_overcurrent_device_subtype, {table}.gs_is_feeder_bay""".format(idcolumn=idcolumn,table=table)))
                cursor.execute(sqlcommand)
                rows = cursor.fetchall()
                totalrows = sumrows(table)
                if not rows:
                    cell_style = 'Normal'
                    cell_alignment = 'False'
                    c = "All are populated correctly."
                else:
                    percent = int(round(int(len(rows))/int(totalrows)*100))
                    cell_style = 'Bad'
                    c = str(len(rows)) + " (" + str(percent) + "%)" + " incorrect values."
                    r = r""
                    sqlcommand = r.join(("SELECT {table}.OBJECTID, {table}.{idcolumn}".format(idcolumn=idcolumn,table=table)))
                    if (facility_id == True and (not(idcolumn == "gs_facility_id"))):
                        sqlcommand = r.join((sqlcommand, """,{table}.gs_facility_id, {table}.gs_overcurrent_device_subtype, {table}.gs_is_feeder_bay
                            FROM {table} 
                            WHERE ({table}.{idcolumn} = '4' AND ({table}.gs_is_feeder_bay <> '1' OR {table}.gs_is_feeder_bay IS NULL)) OR ({table}.gs_overcurrent_device_subtype <> '4' AND {table}.gs_is_feeder_bay <> '0') 
                            GROUP BY {table}.{idcolumn}, {table}.OBJECTID, {table}.gs_facility_id, {table}.gs_overcurrent_device_subtype, {table}.gs_is_feeder_bay""".format(idcolumn=idcolumn,table=table)))
                    else:
                        sqlcommand = r.join((sqlcommand, """, {table}.gs_overcurrent_device_subtype, {table}.gs_is_feeder_bay
                            FROM {table} 
                            WHERE ({table}.{idcolumn} = '4' AND ({table}.gs_is_feeder_bay <> '1' OR {table}.gs_is_feeder_bay IS NULL)) OR ({table}.gs_overcurrent_device_subtype <> '4' AND {table}.gs_is_feeder_bay <> '0') 
                            GROUP BY {table}.{idcolumn}, {table}.OBJECTID, {table}.gs_overcurrent_device_subtype, {table}.gs_is_feeder_bay""".format(idcolumn=idcolumn,table=table)))
                    df = pd.read_sql_query(sqlcommand, conn)
                    sheetname = category + "-FeederBay"
                    createsheet(sheetname)
                    writedf(df, sheetname)
                    write_hyperlink(sheetname)
                break
            else:
                cell_style = 'Neutral'
                c = "This field needs to be added to the database."  
    return c

def excel(xlanalysisfile):
    import openpyxl
    import pyodbc
    import os
    import pandas as pd
    from datetime import date
    from pandas import ExcelWriter
    from pandas import ExcelFile
    from openpyxl.styles import Alignment
    from openpyxl.utils.dataframe import dataframe_to_rows
    global conn
    global category
    global wb
    global ws
    global cell
    global cell_alignment
    global cell_style
    global idcolumn
    global idcolumn1
    global idcolumn2
    global idcolumn3
    conn = pyodbc.connect("Driver={driver};Server=.\SQLEXPRESS;Database={database};Trusted_Connection=yes".format(driver = "{SQL Server}",database = "gs" + number), autocommit = True)
    conn.timeout = 60
    #conn.setencoding('utf-8')  # (Python 3.x syntax)
    #conn.setdecoding(pyodbc.SQL_CHAR, encoding='utf-8')
    #conn.setdecoding(pyodbc.SQL_WCHAR, encoding='utf-8')
    #conn.setencoding(encoding='utf-8')
    global cursor
    cursor = conn.cursor()

    cell_style = 'Normal'
    cell_alignment = 'False'
    cell = 2

    wb = openpyxl.load_workbook(xlanalysisfile, keep_vba=True)
    ws = wb["Overview"]

    c = "OA Data Analysis Summary\n" + number
    ws['A1'] = c
    c = "Review of Snapshot from " + str(date.today())
    ws['C1'] = c

    category = "SUB"
    copysubstation = 0
    table = "gs_electric_station"
    idcolumn = "gs_name"
    c = "Substations (" + str(sumrows (table)) + ")"
    ws['A'+ str(cell)] = c
    cell += 1
    c = uniqueid(idcolumn,table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)
    cell += 1
    idcolumn1 = "gs_facility_id"
    c = duplicateid(idcolumn, idcolumn1, table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)
    idcolumn = "gs_phase"
    cell += 1
    c = fieldsummary(idcolumn,table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)
    idcolumn = "gs_rated_voltage"
    cell += 1
    c = fieldsummary(idcolumn,table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)
    idcolumn = "gs_connection_code"
    cell += 1
    c = fieldsummary(idcolumn,table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)
    idcolumn = "gs_positive_r"
    cell += 1
    c = fieldsummary(idcolumn,table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)
    idcolumn = "gs_positive_x"
    cell += 1
    c = fieldsummary(idcolumn,table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)
    idcolumn = "gs_zero_r"
    cell += 1
    c = fieldsummary(idcolumn,table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)
    idcolumn = "gs_zero_x"
    cell += 1
    c = fieldsummary(idcolumn,table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)
    idcolumn = "gs_bus_voltage"
    cell += 1
    c = fieldsummary(idcolumn,table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)

    category = "CAP"
    table = "gs_capacitor_bank"
    idcolumn = "gs_equipment_location"
    c = "Capacitors (" + str(sumrows (table)) + ")"
    cell += 2
    ws['A'+ str(cell)] = c
    cell += 1
    c = uniqueid(idcolumn,table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)
    idcolumn = "gs_equipment_location"
    idcolumn1 = "gs_facility_id"
    cell += 1
    c = duplicateid(idcolumn, idcolumn1, table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)
    idcolumn = "gs_phase"
    cell += 1
    c = fieldsummary(idcolumn,table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)
    idcolumn = "gs_unit_size_kvar"
    cell += 1
    c = fieldsummary(idcolumn,table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)
    idcolumn = "gs_voltage_rating"
    cell += 1
    c = fieldsummary(idcolumn,table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)
    idcolumn = "gs_status_code"
    cell += 1
    c = fieldsummary(idcolumn,table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)
    idcolumn = "gs_type_code"
    cell += 1
    c = fieldsummary(idcolumn,table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)
    idcolumn = "gs_connection"
    cell += 1
    c = fieldsummary(idcolumn,table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)
    idcolumn = "gs_control_element"
    idcolumn1 = "gs_type_code"
    value = 0
    cell += 1
    c = fieldsummarygt(idcolumn, idcolumn1, value, table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)
    idcolumn = "gs_gang_controlled"
    cell += 1
    c = fieldsummary(idcolumn,table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)
    idcolumn = "gs_control_phase"
    idcolumn1 = "gs_gang_controlled"
    value = 1
    cell += 1
    c = fieldsummaryeq(idcolumn, idcolumn1, value, table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)
    idcolumn = "gs_on_set"
    idcolumn1 = "gs_type_code"
    value = 0
    cell += 1
    c = fieldsummarygt(idcolumn, idcolumn1, value, table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)
    idcolumn = "gs_off_set"
    idcolumn1 = "gs_type_code"
    value = 0
    cell += 1
    c = fieldsummarygt(idcolumn, idcolumn1, value, table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)
    idcolumn = "gs_on_set_winter"
    idcolumn1 = "gs_type_code"
    value = 2
    cell += 1
    c = fieldsummaryeq(idcolumn, idcolumn1, value, table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)
    idcolumn = "gs_off_set_winter"
    idcolumn1 = "gs_type_code"
    value = 2
    cell += 1
    c = fieldsummaryeq(idcolumn, idcolumn1, value, table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)
    idcolumn = "gs_on_month"
    idcolumn1 = "gs_type_code"
    value = 4
    cell += 1
    c = fieldsummarygt(idcolumn, idcolumn1, value, table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)
    idcolumn = "gs_on_day"
    idcolumn1 = "gs_type_code"
    value = 4
    cell += 1
    c = fieldsummarygt(idcolumn, idcolumn1, value, table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)
    idcolumn = "gs_cust_volts_override"
    idcolumn1 = "gs_type_code"
    value = 5
    cell += 1
    c = fieldsummarygt(idcolumn, idcolumn1, value, table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)
    idcolumn = "gs_min_volts_override"
    cell += 1
    c = fieldsummary(idcolumn,table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)
    idcolumn = "gs_max_volts_override"
    cell += 1
    c = fieldsummary(idcolumn,table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)

    category = "OCD"
    table = "gs_overcurrent_device"
    idcolumn = "gs_equipment_location"
    c = "Overcurrent Devices (" + str(sumrows (table)) + ")"
    cell += 2
    ws['A'+ str(cell)] = c
    cell += 1
    c = uniqueid(idcolumn,table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)
    idcolumn = "gs_equipment_location"
    idcolumn1 = "gs_phase"
    cell += 1
    c = duplicateid(idcolumn, idcolumn1, table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)
    idcolumn = "gs_phase"
    cell += 1
    c = fieldsummary(idcolumn,table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)
    idcolumn = "gs_phase"
    idcolumn1 = "gs_device_desc_a"
    idcolumn2 = "gs_device_desc_b"
    idcolumn3 = "gs_device_desc_c"
    cell += 1
    c = nullabc(idcolumn, idcolumn1, idcolumn2, idcolumn3, table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)
    idcolumn = "gs_overcurrent_device_subtype"
    cell += 1
    c = fieldsummary(idcolumn,table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)
    idcolumn = "gs_switch_description"
    idcolumn1 = "gs_overcurrent_device_subtype"
    value = 1
    cell += 1
    c = fieldsummaryeq(idcolumn,idcolumn1, value, table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)
    idcolumn = "gs_equipment_location"
    cell += 1
    c = is_feeder_bay(idcolumn,table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)

    category = "SWI"
    table = "gs_switch"
    idcolumn = "gs_equipment_location"
    c = "Switches (" + str(sumrows (table)) + ")"
    cell += 3
    ws['A'+ str(cell)] = c
    cell += 1
    c = uniqueid(idcolumn,table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)
    idcolumn = "gs_equipment_location"
    idcolumn1 = "gs_facility_id"
    cell += 1
    c = duplicateid(idcolumn, idcolumn1, table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)
    idcolumn = "gs_phase"
    cell += 1
    c = fieldsummary(idcolumn,table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)
    idcolumn = "gs_switch_status"
    cell += 1
    c = fieldsummary(idcolumn,table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)
    idcolumn = "gs_switch_description"
    cell += 1
    c = fieldsummary(idcolumn,table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)

    category = "XFMR"
    table = "gs_transformer"
    idcolumn = "gs_equipment_location"
    c = "Transformers (" + str(sumrows (table)) + ")"
    cell += 2
    ws['A'+ str(cell)] = c
    cell += 1
    c = uniqueid(idcolumn,table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)
    idcolumn = "gs_equipment_location"
    idcolumn1 = "gs_bank_id"
    cell += 1
    c = duplicate_xfmr(idcolumn, idcolumn1, table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)
    idcolumn = "gs_phase"
    cell += 1
    c = fieldsummary(idcolumn,table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)
    idcolumn = "gs_xfmr_conductor_description"
    cell += 1
    c = fieldsummary(idcolumn,table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)
    idcolumn = "gs_phase"
    idcolumn1 = "gs_tran_kva_a"
    idcolumn2 = "gs_tran_kva_b"
    idcolumn3 = "gs_tran_kva_c"
    cell += 1
    c = nullabc(idcolumn, idcolumn1, idcolumn2, idcolumn3, table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)
    idcolumn = "gs_winding_connection"
    cell += 1
    c = fieldsummary(idcolumn,table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)
    idcolumn = "gs_rated_input_voltage"
    cell += 1
    c = fieldsummary(idcolumn,table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)
    idcolumn = "gs_rated_input_voltage"
    idcolumn1 = "gs_rated_output_voltage"
    xfmr_voltage(idcolumn,idcolumn1,table)
    idcolumn = "gs_rated_output_voltage"
    cell += 1
    c = fieldsummary(idcolumn,table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)
    #idcolumn = "gs_equipment_location"
    #idcolumn1 = "gs_bank_id"
    #c = duplicateid(idcolumn, idcolumn1, table) + " They should have gs_bank_ids added if they are banked."
    #ws['C45'] = c
    #ws['C45'].style = cell_style
    #ws['C45'].alignment = Alignment(wrap_text=cell_alignment)
    idcolumn = "gs_is_substation_transformer"
    cell += 2
    c = fieldsummary(idcolumn,table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)
    idcolumn = "gs_positive_r"
    idcolumn1 = "gs_is_substation_transformer"
    value = "\'true\'"
    cell += 1
    c = fieldsummaryeqtext(idcolumn, idcolumn1, value, table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)
    idcolumn = "gs_positive_x"
    idcolumn1 = "gs_is_substation_transformer"
    value = "\'true\'"
    cell += 1
    c = fieldsummaryeqtext(idcolumn, idcolumn1, value, table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)
    idcolumn = "gs_zero_r"
    idcolumn1 = "gs_is_substation_transformer"
    value = "\'true\'"
    cell += 1
    c = fieldsummaryeqtext(idcolumn, idcolumn1, value, table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)
    idcolumn = "gs_zero_x"
    idcolumn1 = "gs_is_substation_transformer"
    value = "\'true\'"
    cell += 1
    c = fieldsummaryeqtext(idcolumn, idcolumn1, value, table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)
    idcolumn = "gs_impedance"
    idcolumn1 = "gs_is_substation_transformer"
    value = "\'true\'"
    cell += 1
    c = fieldsummaryeqtext(idcolumn, idcolumn1, value, table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)
    idcolumn = "gs_impedance_angle"
    idcolumn1 = "gs_is_substation_transformer"
    value = "\'true\'"
    cell += 1
    c = fieldsummaryeqtext(idcolumn, idcolumn1, value, table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)
    idcolumn = "gs_is_center_tap"
    idcolumn1 = "gs_is_substation_transformer"
    value = "\'true\'"
    cell += 1
    c = fieldsummaryeqtext(idcolumn, idcolumn1, value, table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)

    category = "VREG"
    table = "gs_voltage_regulator"
    idcolumn = "gs_equipment_location"
    c = "Voltage Regulators (" + str(sumrows (table)) + ")"
    cell += 2
    ws['A'+ str(cell)] = c
    cell += 1
    c = uniqueid(idcolumn,table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)
    idcolumn = "gs_equipment_location"
    idcolumn1 = "gs_facility_id"
    cell += 1
    c = duplicateid(idcolumn, idcolumn1, table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)
    idcolumn = "gs_phase"
    cell += 1
    c = fieldsummary(idcolumn,table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)
    idcolumn = "gs_phase"
    idcolumn1 = "gs_regulator_a"
    idcolumn2 = "gs_regulator_b"
    idcolumn3 = "gs_regulator_c"
    cell += 1
    c = nullabc(idcolumn, idcolumn1, idcolumn2, idcolumn3, table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)
    idcolumn = "gs_winding_connection"
    cell += 1
    c = fieldsummary(idcolumn,table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)
    idcolumn = "gs_nominal_voltage"
    cell += 1
    c = fieldsummary(idcolumn,table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)
    idcolumn = "gs_base_volts"
    cell += 1
    c = fieldsummary(idcolumn,table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)
    idcolumn = "gs_bandwidth"
    cell += 1
    c = fieldsummary(idcolumn,table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)
    idcolumn = "gs_ldc_a_total"
    cell += 1
    c = fieldsummary(idcolumn,table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)
    idcolumn = "gs_ldc_r_total"
    cell += 1
    c = fieldsummary(idcolumn,table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)
    idcolumn = "gs_ldc_x_total"
    cell += 1
    c = fieldsummary(idcolumn,table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)
    idcolumn = "gs_regulator_mode"
    cell += 1
    c = fieldsummary(idcolumn,table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)
    idcolumn = "gs_step_a"
    cell += 1
    c = fieldsummary(idcolumn,table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)
    idcolumn = "gs_step_b"
    cell += 1
    c = fieldsummary(idcolumn,table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)
    idcolumn = "gs_step_c"
    cell += 1
    c = fieldsummary(idcolumn,table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)
    idcolumn = "gs_block_step"
    cell += 1
    c = fieldsummary(idcolumn,table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)
    idcolumn = "gs_regulator_type"
    cell += 1
    c = fieldsummary(idcolumn,table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)
    idcolumn = "gs_controlling_phase"
    idcolumn1 = "gs_regulator_type"
    value = 1
    cell += 1
    c = fieldsummaryeq(idcolumn, idcolumn1, value, table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)
    idcolumn = "gs_regulating_phase"
    cell += 1
    c = fieldsummary(idcolumn,table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)

    category = "SPAN"
    table = "gs_span"
    c = "Conductors (" + str(sumrows (table)) + ")"
    cell += 2
    ws['A'+ str(cell)] = c
    idcolumn = "gs_phase"
    cell += 1
    c = fieldsummary(idcolumn,table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)
    idcolumn = "gs_phase"
    idcolumn1 = "gs_conductor_a"
    idcolumn2 = "gs_conductor_b"
    idcolumn3 = "gs_conductor_c"
    cell += 1
    c = nullabc(idcolumn, idcolumn1, idcolumn2, idcolumn3, table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)
    idcolumn = "gs_conductor_n"
    cell += 1
    c = neutral(idcolumn,table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)
    idcolumn = "gs_subtype_cd"
    cell += 1
    c = fieldsummary(idcolumn,table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=True)
    idcolumn = "gs_construction_desc"
    cell += 1
    c = missingfield(idcolumn,table)
    ws['C'+ str(cell)] = c + " This is common in GIS data and we compensate by populating that field with RUS standards for the most common configurations."
    ws['C'+ str(cell)].style = "Normal"
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)
    ws.column_dimensions['C'].auto_size = True
    cell += 1
    span_assembly(table)

    category = "MTR"
    table = "gs_motor"
    c = "Motors (" + str(sumrows (table)) + ")"
    cell += 2
    ws['A'+ str(cell)] = c
    idcolumn = "gs_equipment_location"
    cell += 1
    c = uniqueid(idcolumn,table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)
    idcolumn = "gs_nema_type"
    cell += 1
    c = fieldsummary(idcolumn,table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)
    idcolumn = "gs_soft_start_tap"
    cell += 1
    c = fieldsummary(idcolumn,table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)
    idcolumn = "gs_rated_hp"
    cell += 1
    c = fieldsummary(idcolumn,table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)
    idcolumn = "gs_power_factor"
    cell += 1
    c = fieldsummary(idcolumn,table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)
    idcolumn = "gs_locked_rotor_multiplier"
    cell += 1
    c = fieldsummary(idcolumn,table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)
    idcolumn = "gs_locked_rotor_pf"
    cell += 1
    c = fieldsummary(idcolumn,table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)

    category = "GEN"
    table = "gs_generator"
    c = "Generators (" + str(sumrows (table)) + ")"
    cell += 2
    ws['A'+ str(cell)] = c
    idcolumn = "gs_subtype_cd"
    cell += 1
    c = fieldsummary(idcolumn,table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)
    idcolumn = "gs_max_kw_out"
    cell += 1
    c = fieldsummary(idcolumn,table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)
    idcolumn = "gs_rated_kva"
    cell += 1
    c = fieldsummary(idcolumn,table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)
    idcolumn = "gs_max_kvar_lagg"
    cell += 1
    c = fieldsummary(idcolumn,table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)
    idcolumn = "gs_power_factor"
    cell += 1
    c = fieldsummary(idcolumn,table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)
    idcolumn = "gs_on_off"
    cell += 1
    c = fieldsummary(idcolumn,table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)
    idcolumn = "gs_installation_date"
    cell += 1
    #c = fieldsummary(idcolumn,table)
    #ws['C'+ str(cell)] = c
    #ws['C'+ str(cell)].style = cell_style
    #ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)
    idcolumn = "gs_fault_contribution"
    cell += 1
    c = fieldsummary(idcolumn,table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)
    idcolumn = "gs_inverter_efficiency"
    cell += 1
    c = fieldsummary(idcolumn,table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)
    idcolumn = "gs_curtailing_component_id"
    cell += 1
    c = fieldsummary(idcolumn,table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)
    idcolumn = "gs_power_factor_response"
    cell += 1
    c = fieldsummary(idcolumn,table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)
    idcolumn = "gs_rpm"
    cell += 1
    c = fieldsummary(idcolumn,table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)
    idcolumn = "gs_mva_base"
    cell += 1
    c = fieldsummary(idcolumn,table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)
    idcolumn = "gs_kva_base"
    cell += 1
    c = fieldsummary(idcolumn,table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)
    idcolumn = "gs_positive_sequence_reactance"
    cell += 1
    c = fieldsummary(idcolumn,table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)
    idcolumn = "gs_saturated_sequence_reactance"
    cell += 1
    c = fieldsummary(idcolumn,table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)
    idcolumn = "gs_transient_reactance"
    cell += 1
    c = fieldsummary(idcolumn,table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)
    idcolumn = "gs_subtransient_reactance"
    cell += 1
    c = fieldsummary(idcolumn,table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)
    idcolumn = "gs_transient_time_constant"
    cell += 1
    c = fieldsummary(idcolumn,table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)
    idcolumn = "gs_subtransient_time_constant"
    cell += 1
    c = fieldsummary(idcolumn,table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)
    idcolumn = "gs_num_poles"
    cell += 1
    c = fieldsummary(idcolumn,table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)
    idcolumn = "gs_slip"
    cell += 1
    c = fieldsummary(idcolumn,table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)
    idcolumn = "gs_stator_resistance"
    cell += 1
    c = fieldsummary(idcolumn,table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)
    idcolumn = "gs_stator_reactance"
    cell += 1
    c = fieldsummary(idcolumn,table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)
    idcolumn = "gs_rotor_resistance"
    cell += 1
    c = fieldsummary(idcolumn,table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)
    idcolumn = "gs_rotor_reactance"
    cell += 1
    c = fieldsummary(idcolumn,table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)
    idcolumn = "gs_magnetizing_reactance"
    cell += 1
    c = fieldsummary(idcolumn,table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)
    idcolumn = "gs_crowbar_resistance"
    cell += 1
    c = fieldsummary(idcolumn,table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)
    idcolumn = "gs_tilt"
    cell += 1
    c = fieldsummary(idcolumn,table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)
    idcolumn = "gs_azimuth"
    cell += 1
    c = fieldsummary(idcolumn,table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)
    idcolumn = "gs_cpr_site_id"
    cell += 1
    c = fieldsummary(idcolumn,table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)
    idcolumn = "gs_charge"
    cell += 1
    c = fieldsummary(idcolumn,table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)
    idcolumn = "gs_kw"
    cell += 1
    c = fieldsummary(idcolumn,table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)
    idcolumn = "gs_cell_voltage"
    cell += 1
    c = fieldsummary(idcolumn,table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)
    idcolumn = "gs_cell_resistance"
    cell += 1
    c = fieldsummary(idcolumn,table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)
    idcolumn = "gs_cell_count"
    cell += 1
    c = fieldsummary(idcolumn,table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)
    idcolumn = "gs_blade_diameter"
    cell += 1
    c = fieldsummary(idcolumn,table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)
    idcolumn = "gs_performance_coefficient"
    cell += 1
    c = fieldsummary(idcolumn,table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)
    idcolumn = "gs_generator_efficiency"
    cell += 1
    c = fieldsummary(idcolumn,table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)
    idcolumn = "gs_gear_box_efficiency"
    cell += 1
    c = fieldsummary(idcolumn,table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)

    category = "SERV"
    table = "gs_service_point"
    c = "Service Points (" + str(sumrows (table)) + ")"
    cell += 2
    ws['A'+ str(cell)] = c
    cell += 1
    idcolumn = "gs_service_map_location"
    c = uniqueid(idcolumn,table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)
    idcolumn1 = "gs_service_number"
    cell += 1
    c = duplicateid(idcolumn, idcolumn1, table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)
    idcolumn = "gs_phase"
    cell += 1
    c = fieldsummary(idcolumn,table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)
    cell += 1
    c = duplicate_system()
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)

    print("Done.")
    conn.close()
    wb.save(xlanalysisfile)
    wb.close()
    return

from flask import send_from_directory

@app.route('/uploads/<filename>')
def uploaded_file(filename):
    return send_from_directory(app.config['UPLOAD_FOLDER'],
                               filename)