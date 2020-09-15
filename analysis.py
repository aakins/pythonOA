import excel

def table_exists(table):
    sqlcommand = r"""SELECT COUNT(*)
                FROM information_schema.tables
                WHERE table_name = '{table}';
                """.format(table=table)
    excel.cursor.execute(sqlcommand)
    if excel.cursor.fetchone()[0] == 1:
        return True
    return False

def createsheet(sheetname):
    if sheetname in excel.wb.sheetnames:
        #excel.wb.remove(wb[sheetname])
        #excel.wb.create_sheet(sheetname)
        return
    else:
        excel.wb.create_sheet(sheetname)
    return
    
def writedf(df, sheetname):
    from openpyxl.utils.dataframe import dataframe_to_rows
    excel.ws = excel.wb[sheetname]
    for r in dataframe_to_rows(df, index=False, header=True):
        excel.ws.append(r)
    for cell in excel.ws[1]:
        cell.style = '40 % - Accent3'
    #ws.column_dimensions['A'].auto_size = True
    #ws.column_dimensions['B'].auto_size = True
    #ws.column_dimensions['C'].auto_size = True
    #ws.column_dimensions['D'].auto_size = True
    createsheet("Tasklist")
    excel.ws = excel.wb["Tasklist"]
    excel.ws.append(tuple([sheetname]))
    excel.ws = excel.wb["Overview"]
    return
    
def write_hyperlink(sheetname):
    hyperlink = "#'" + sheetname + "'!A1"
    excel.ws['D'+ str(excel.cell)].hyperlink = hyperlink
    excel.ws['D'+ str(excel.cell)].value = sheetname
    excel.ws['D'+ str(excel.cell)].style = "Hyperlink"
    return

def collapse_rows(start, end):
    excel.ws = excel.wb["Overview"]
    excel.ws.row_dimensions.group(start, end, hidden=True)
    return

def sumrows(table):
    try:
        excel.cursor.execute("SELECT COUNT(*) FROM {table}".format(table=table))
        row = excel.cursor.fetchone()
        totalrows = str(row[0])
        return totalrows
    except:
        return
    
def uniqueid(idcolumn, table):
    import pandas as pd
    c = "Error."
    facility_id = False
    for row in excel.cursor.columns(table=table):
        if "gs_facility_id" in row:
            facility_id = True
    #global excel.cell_style 
    excel.cell_style = 'Normal'
    global cell_alignment
    cell_alignment = 'False'
    rows = excel.cursor.columns(table=table)
    if int(sumrows(table)) == 0:
                excel.cell_style = 'Bad'
                c = "This table is empty."
    else:
        for row in excel.cursor.columns(table=table):
            if idcolumn in row:
                r = r""
                sqlcommand = r.join(("SELECT {table}.OBJECTID, {table}.{idcolumn}".format(idcolumn=idcolumn,table=table)))
                if (facility_id == True and (not(idcolumn == "gs_facility_id"))):
                    sqlcommand = r.join((sqlcommand, """,{table}.gs_facility_id, Count({table}.OBJECTID) AS [Count] 
                        FROM {table} WHERE ({table}.{idcolumn} IS null OR {table}.{idcolumn} = '') 
                        GROUP BY {table}.{idcolumn}, {table}.OBJECTID""".format(idcolumn=idcolumn,table=table)))
                else:
                    sqlcommand = r.join((sqlcommand, """, Count({table}.OBJECTID) AS [Count] 
                        FROM {table} WHERE ({table}.{idcolumn} IS null OR {table}.{idcolumn} = '') 
                        GROUP BY {table}.{idcolumn}, {table}.OBJECTID""".format(idcolumn=idcolumn,table=table)))
                if (facility_id == True and (not(idcolumn == "gs_facility_id"))):
                    sqlcommand = r.join((sqlcommand, """,{table}.gs_facility_id 
                        ORDER BY {table}.{idcolumn}, Count({table}.OBJECTID) DESC, {table}.OBJECTID
                        """.format(idcolumn=idcolumn,table=table)))
                else:
                    sqlcommand = r.join((sqlcommand, """ 
                        ORDER BY {table}.{idcolumn}, Count({table}.OBJECTID) DESC, {table}.OBJECTID
                        """.format(idcolumn=idcolumn,table=table)))
                excel.cursor.execute(sqlcommand)
                rows = excel.cursor.fetchall()
                totalrows = sumrows(table)
                if not rows:
                    excel.cell_style = 'Normal'
                    cell_alignment = 'False'
                    c = "All IDs are populated."
                else:
                    if (facility_id == True and (not(idcolumn == "gs_facility_id"))):
                        countid = str(sum(e[3] for e in rows))
                    else:
                        countid = str(sum(e[2] for e in rows))
                    percent = int(round(int(countid)/int(totalrows)*100))
                    excel.cell_style = 'Bad'
                    c = str(countid) + " (" + str(percent) + "%)" + " blank values."
                    r = r""
                    sqlcommand = r.join(("SELECT {table}.OBJECTID, {table}.{idcolumn}".format(idcolumn=idcolumn,table=table)))
                    if (facility_id == True and (not(idcolumn == "gs_facility_id"))):
                        sqlcommand = r.join((sqlcommand, """,{table}.gs_facility_id, Count({table}.OBJECTID) AS [Count] 
                            FROM {table} WHERE ({table}.{idcolumn} IS null OR {table}.{idcolumn} = '') 
                            GROUP BY {table}.{idcolumn}, {table}.OBJECTID""".format(idcolumn=idcolumn,table=table)))
                    else:
                        sqlcommand = r.join((sqlcommand, """, Count({table}.OBJECTID) AS [Count] 
                            FROM {table} WHERE ({table}.{idcolumn} IS null OR {table}.{idcolumn} = '') 
                            GROUP BY {table}.{idcolumn}, {table}.OBJECTID""".format(idcolumn=idcolumn,table=table)))
                    if (facility_id == True and (not(idcolumn == "gs_facility_id"))):
                        sqlcommand = r.join((sqlcommand, """,{table}.gs_facility_id 
                            ORDER BY {table}.{idcolumn}, Count({table}.OBJECTID) DESC, {table}.OBJECTID
                            """.format(idcolumn=idcolumn,table=table)))
                    else:
                        sqlcommand = r.join((sqlcommand, """ 
                            ORDER BY {table}.{idcolumn}, Count({table}.OBJECTID) DESC, {table}.OBJECTID
                            """.format(idcolumn=idcolumn,table=table)))
                    df = pd.read_sql_query(sqlcommand, excel.conn)
                    sheetname = excel.category + "-Blanks"
                    createsheet(sheetname)
                    writedf(df, sheetname)
                    write_hyperlink(sheetname)
                break
            else:
                excel.cell_style = 'Neutral'
                c = "This field needs to be added to the database."  
    return c
    
def duplicateid(idcolumn, idcolumn1, table):
    import pandas as pd
    c = "Error."
    #global excel.cell_style
    excel.cell_style = 'Normal'
    global cell_alignment
    cell_alignment = 'False'
    #rows = excel.cursor.columns(table=table)
    for row in excel.cursor.columns(table=table):
        if "gs_facility_id" in row:
            facility_id = True
        else:
            facility_id = False
    for row in excel.cursor.columns(table=table):
        if "gs_bank_id" in row:
            bank_id = True
        else:
            bank_id = False
    if int(sumrows(table)) == 0:
        excel.cell_style = 'Bad'
        c = "This table is empty."
    else:
        for row in excel.cursor.columns(table=table):
            if idcolumn in row:
                for row in excel.cursor.columns(table=table):
                    if idcolumn1 in row:
                        sqlcommand = r"""SELECT DISTINCT {table}.{idcolumn1}, {table}.{idcolumn}
                                        FROM {table}
                                        WHERE ((({table}.{idcolumn}) In (SELECT [{idcolumn}] FROM [{table}] As Tmp GROUP BY [{idcolumn}] HAVING Count(*)>1 )))
                                        ORDER BY {table}.{idcolumn};
                                    """.format(idcolumn=idcolumn,idcolumn1=idcolumn1,table=table)
                        excel.cursor.execute(sqlcommand)
                        row = excel.cursor.fetchall()
                        if not row:
                            excel.cell_style = 'Normal'
                            cell_alignment = 'False'
                            c = "No duplicates found."
                        else:
                            excel.cell_style = 'Bad'
                            c = str(len(row)) + " duplicates exist."
                            sqlcommand = r"""SELECT DISTINCT {table}.OBJECTID, {table}.{idcolumn}, {table}.{idcolumn1}
                                    FROM {table}
                                    WHERE ((({table}.{idcolumn}) In (SELECT [{idcolumn}] FROM [{table}] As Tmp GROUP BY [{idcolumn}] HAVING Count(*)>1 )))
                                    GROUP BY {table}.OBJECTID, {table}.{idcolumn}, {table}.{idcolumn1}
                                    ORDER BY {table}.{idcolumn}, {table}.OBJECTID, {table}.{idcolumn1}
                                    """.format(idcolumn=idcolumn,idcolumn1=idcolumn1,table=table)
                            r = r""
                            sqlcommand = r.join(("SELECT DISTINCT {table}.OBJECTID, {table}.{idcolumn}, {table}.{idcolumn1}".format(idcolumn=idcolumn,idcolumn1=idcolumn1,table=table)))
                            if (facility_id == True and (not(idcolumn == "gs_facility_id")) and bank_id == True):
                                sqlcommand = r.join((sqlcommand, """, {table}.gs_facility_id 
                                    FROM {table} 
                                    WHERE ((({table}.{idcolumn}) 
                                        In (SELECT [{idcolumn}] 
                                        FROM [{table}] As Tmp 
                                        GROUP BY [{idcolumn}] HAVING Count(*)>1 )))
                                    GROUP BY {table}.OBJECTID, {table}.{idcolumn}, {table}.{idcolumn1}, {table}.gs_facility_id, {table}.gs_bank_id
                                    ORDER BY {table}.{idcolumn}, {table}.gs_bank_id, {table}.{idcolumn1}, {table}.gs_facility_id, {table}.OBJECTID  """.format(idcolumn=idcolumn,idcolumn1=idcolumn1,table=table)))
                            else:
                                sqlcommand = r.join((sqlcommand, """ FROM {table} 
                                    WHERE ((({table}.{idcolumn}) 
                                        In (SELECT [{idcolumn}] 
                                        FROM [{table}] As Tmp 
                                        GROUP BY [{idcolumn}] HAVING Count(*)>1 )))
                                    GROUP BY {table}.OBJECTID, {table}.{idcolumn}, {table}.{idcolumn1}
                                    ORDER BY {table}.OBJECTID, {table}.{idcolumn}, {table}.{idcolumn1}""".format(idcolumn=idcolumn,idcolumn1=idcolumn1,table=table)))
                            df = pd.read_sql_query(sqlcommand, excel.conn)
                            sheetname = excel.category + "-" + idcolumn[len("gs_"):]
                            createsheet(sheetname)
                            writedf(df, sheetname)
                            write_hyperlink(sheetname)
                        break
                    else:
                        excel.cell_style = 'Neutral'
                        c = "The field ''" + str(idcolumn1) + "' needs to be added to the database."
            else:
                excel.cell_style = 'Neutral'
                c = "The field ''" + str(idcolumn) + "' needs to be added to the database."
    return c

def fieldsummary(idcolumn, table):
    import pandas as pd
    c = "Error."
    #global excel.cell_style 
    excel.cell_style = 'Normal'
    global cell_alignment
    cell_alignment = 'False'
    facility_id = False
    for row in excel.cursor.columns(table=table):
        if "gs_facility_id" in row:
            facility_id = True
    for row in excel.cursor.columns(table=table):
        if "gs_equipment_location" in row:
            equipment_location = True
    for row in excel.cursor.columns(table=table):
        if idcolumn in row:
            file_list = []
            if idcolumn == "gs_phase":
                sqlcommand = r"""SELECT {table}.{idcolumn}, Count({table}.OBJECTID) AS [Count]
                            FROM {table}
                            GROUP BY {table}.{idcolumn}
                            ORDER BY CASE WHEN {idcolumn} = 'NULL' THEN '1'
                                          WHEN {idcolumn} = 'UNK' THEN '2'
                                          ELSE {idcolumn} END ASC
                            """.format(idcolumn=idcolumn,table=table)
            elif idcolumn == "gs_installation_date":
                sqlcommand = r"""SELECT {table}.{idcolumn}, Count({table}.OBJECTID) AS [Count]
                            FROM {table}
                            GROUP BY {table}.{idcolumn}
                            """.format(idcolumn=idcolumn,table=table)
            else:
                sqlcommand = r"""SELECT {table}.{idcolumn}, Count({table}.OBJECTID) AS [Count]
                            FROM {table}
                            GROUP BY {table}.{idcolumn}
                            ORDER BY CASE WHEN {idcolumn} = TRY_CONVERT(numeric, 'NULL') THEN '1'
                                          WHEN {idcolumn} = TRY_CONVERT(numeric, 'UNK') THEN '2'
                                          ELSE {idcolumn} END ASC
                            """.format(idcolumn=idcolumn,table=table)
            #print(idcolumn + " " + sqlcommand)
            excel.cursor.execute(sqlcommand)
            row = excel.cursor.fetchall()
            totalrows = sumrows(table)
            if not row:
                excel.cell_style = 'Bad'
                c = "This table is empty."
            else:
                if int(sumrows(table)) == int(row[0][1]):
                    excel.cell_style = 'Normal'
                    cell_alignment = 'False'
                    if row[0][0] == None:
                        c = "All fields are NULL."
                    elif row[0][0] == 0:
                        c = "All fields are populated with '0'."
                    else:
                        c = "All fields are populated with '" + str(row[0][0]) +"'."
                    if (row[0][0] == None) or (row[0][0] == "") or (row[0][0] == "UNK")or (row[0] == "0E-8"):
                            excel.cell_style = 'Bad'
                            r = r""
                            sqlcommand = r.join(("SELECT {table}.OBJECTID, {table}.{idcolumn}".format(idcolumn=idcolumn,table=table)))
                            sqlcommand = r.join((sqlcommand, """ FROM {table}
                                 WHERE ({table}.{idcolumn} IS null 
                                 OR {table}.{idcolumn} = TRY_CONVERT(numeric, '') 
                                 OR {table}.{idcolumn} LIKE '%UNK%')
                                 GROUP BY {table}.OBJECTID, {table}.{idcolumn}
                                 ORDER BY {table}.OBJECTID, {table}.{idcolumn}""".format(idcolumn=idcolumn,table=table)))
                            df = pd.read_sql_query(sqlcommand, excel.conn)
                            sheetname = excel.category + "-" + idcolumn[len("gs_"):]
                            createsheet(sheetname)
                            writedf(df, sheetname)
                            write_hyperlink(sheetname)
                else:
                    excel.cell_style = 'Normal'
                    for row in row:
                        if (row[0] == None) or (row[0] == "") or (row[0] == "UNK")or (row[0] == "0E-8"):
                            excel.cell_style = 'Bad'
                        fcol = row[0]
                        if fcol == 0:
                            fcol = "0"
                        if fcol == None:
                            fcol = "NULL"
                        fnum = row[1]
                        fper = int(round(int(fnum)/int(totalrows)*100))
                        file_list.append({"fcol": fcol, "fnum": fnum, "fper": fper})
                    cell_alignment = 'True'
                    c = r""        
                    c = c.join(("{fnum} ({fper}%) populated with '{fcol}'.\n".format(fcol=fl['fcol'], fnum=fl['fnum'], fper=fl['fper']) for fl in file_list))
                    if excel.cell_style == 'Bad':

                        r = r""
                        sqlcommand = r.join(("SELECT {table}.OBJECTID, {table}.{idcolumn}".format(idcolumn=idcolumn,table=table)))
                        if (facility_id == True and (not(idcolumn == "gs_facility_id")) and equipment_location == True and (not(idcolumn == "gs_equipment_location"))):
                            sqlcommand = r.join((sqlcommand, """,{table}.gs_equipment_location, {table}.gs_facility_id 
                                FROM {table}
                                WHERE ({table}.{idcolumn} IS null 
                                    OR {table}.{idcolumn} = TRY_CONVERT(numeric, '') 
                                    OR {table}.{idcolumn} LIKE '%UNK%')
                                    GROUP BY {table}.OBJECTID, {table}.{idcolumn},{table}.gs_equipment_location, {table}.gs_facility_id
                                    ORDER BY {table}.OBJECTID, {table}.{idcolumn},{table}.gs_equipment_location, {table}.gs_facility_id""".format(idcolumn=idcolumn,table=table)))
                        elif (facility_id == True and (not(idcolumn == "gs_facility_id"))):
                            sqlcommand = r.join((sqlcommand, """,{table}.gs_facility_id 
                                FROM {table}
                                WHERE ({table}.{idcolumn} IS null 
                                    OR {table}.{idcolumn} = TRY_CONVERT(numeric, '') 
                                    OR {table}.{idcolumn} LIKE '%UNK%')
                                    GROUP BY {table}.OBJECTID, {table}.{idcolumn}, {table}.gs_facility_id
                                    ORDER BY {table}.OBJECTID, {table}.{idcolumn}, {table}.gs_facility_id""".format(idcolumn=idcolumn,table=table))) 
                        else:
                            sqlcommand = r.join((sqlcommand, """ FROM {table}
                                WHERE ({table}.{idcolumn} IS null 
                                    OR {table}.{idcolumn} = TRY_CONVERT(numeric, '') 
                                    OR {table}.{idcolumn} LIKE '%UNK%')
                                    GROUP BY {table}.OBJECTID, {table}.{idcolumn}
                                    ORDER BY {table}.OBJECTID, {table}.{idcolumn}""".format(idcolumn=idcolumn,table=table)))
                        df = pd.read_sql_query(sqlcommand, excel.conn)
                        sheetname = excel.category + "-" + idcolumn[len("gs_"):]
                        createsheet(sheetname)
                        writedf(df, sheetname)
                        write_hyperlink(sheetname)
            break
        else:
            excel.cell_style = 'Neutral'
            c = "This field needs to be added to the database."
    
    return c

def fieldsummarygt(idcolumn, idcolumn1, value, table):
    import pandas as pd
    c = "Error."
    #global excel.cell_style 
    excel.cell_style = 'Normal'
    global cell_alignment
    cell_alignment = 'False'
    facility_id = False
    idcolumnFound = False
    idcolumn1Found = False						 						  
    for row in excel.cursor.columns(table=table):
        if "gs_facility_id" in row:
            facility_id = True
    for row in excel.cursor.columns(table=table):
        if idcolumn in row:
            idcolumnFound = True
    for row in excel.cursor.columns(table=table):
        if idcolumn1 in row:
            idcolumn1Found = True
    for row in excel.cursor.columns(table=table):
        if idcolumnFound == True and idcolumn1Found == True:						   
            file_list = []
            sqlcommand = r"""SELECT {table}.{idcolumn}, Count({table}.OBJECTID) AS [Count]
                        FROM {table}
                        WHERE TRY_CONVERT(numeric, {idcolumn1}) > {value}
                        GROUP BY {table}.{idcolumn}
                        ORDER BY CASE WHEN {idcolumn} = TRY_CONVERT(numeric, 'NULL') THEN '1'
                                      WHEN {idcolumn} = TRY_CONVERT(numeric, 'UNK') THEN '2'
                                      ELSE {idcolumn} END ASC
                        """.format(idcolumn=idcolumn,idcolumn1=idcolumn1,value=value,table=table)
            excel.cursor.execute(sqlcommand)
            row = excel.cursor.fetchall()
            totalrows = sumrows(table)
            if not row:
                excel.cell_style = 'Bad'
                c = ""
            else:
                if int(sumrows(table)) == int(row[0][1]):
                    excel.cell_style = 'Normal'
                    cell_alignment = 'False'
                    if row[0][0] == None:
                        c = "All fields are NULL."
                    elif row[0][0] == 0:
                        c = "All fields are populated with '0'."
                    else:
                        c = "All fields are populated with '" + str(row[0][0]) +"'."
                    if (row[0][0] == None) or (row[0][0] == "") or (row[0][0] == "UNK") or (row[0] == "0E-8"):
                            excel.cell_style = 'Bad'
                else:
                    excel.cell_style = 'Normal'
                    for row in row:
                        if (row[0] == None) or (row[0] == "") or (row[0] == "UNK") or (row[0] == "0E-8"):
                            excel.cell_style = 'Bad'
                        fcol = row[0]
                        if fcol == 0:
                            fcol = "0"
                        if fcol == None:
                            fcol = "NULL"
                        fnum = row[1]
                        fper = int(round(int(fnum)/int(totalrows)*100))
                        file_list.append({"fcol": fcol, "fnum": fnum, "fper": fper})
                    cell_alignment = 'True'
                    c = r""        
                    c = c.join(("{fnum} ({fper}%) populated with '{fcol}'.\n".format(fcol=fl['fcol'], fnum=fl['fnum'], fper=fl['fper']) for fl in file_list))
                    if excel.cell_style == 'Bad':

                        r = r""
                        sqlcommand = r.join(("SELECT {table}.OBJECTID, {table}.{idcolumn}".format(idcolumn=idcolumn,table=table)))
                        if (facility_id == True and (not(idcolumn == "gs_facility_id"))):
                            sqlcommand = r.join((sqlcommand, """,{table}.gs_facility_id 
                                FROM {table}
                                WHERE ({table}.{idcolumn} IS null 
                                    OR {table}.{idcolumn} = TRY_CONVERT(numeric, '') 
                                    OR {table}.{idcolumn} LIKE '%UNK%')
                                    GROUP BY {table}.OBJECTID, {table}.{idcolumn}, {table}.gs_facility_id
                                    ORDER BY {table}.OBJECTID, {table}.{idcolumn}, {table}.gs_facility_id""".format(idcolumn=idcolumn,table=table)))
                        else:
                            sqlcommand = r.join((sqlcommand, """ FROM {table}
                                WHERE ({table}.{idcolumn} IS null 
                                    OR {table}.{idcolumn} = TRY_CONVERT(numeric, '') 
                                    OR {table}.{idcolumn} LIKE '%UNK%')
                                    GROUP BY {table}.OBJECTID, {table}.{idcolumn}
                                    ORDER BY {table}.OBJECTID, {table}.{idcolumn}""".format(idcolumn=idcolumn,table=table)))
                        df = pd.read_sql_query(sqlcommand, excel.conn)
                        sheetname = excel.category + "-" + idcolumn[len("gs_"):]
                        createsheet(sheetname)
                        writedf(df, sheetname)
                        write_hyperlink(sheetname)
            break
        else:
            excel.cell_style = 'Neutral'
            if idcolumnFound == False:
                c = idcolumn + " needs to be added to the database."
            elif idcolumn1Found == False:
                c = idcolumn1 + " needs to be added to the database."
    
    return c

def fieldsummaryeq(idcolumn, idcolumn1, value, table):
    import pandas as pd
    c = "Error."
    #global excel.cell_style 
    excel.cell_style = 'Normal'
    global cell_alignment
    cell_alignment = 'False'
    facility_id = False
    idcolumnFound = False
    idcolumn1Found = False				 
    for row in excel.cursor.columns(table=table):
        if "gs_facility_id" in row:
            facility_id = True
    for row in excel.cursor.columns(table=table):
        if idcolumn in row:
            idcolumnFound = True
    for row in excel.cursor.columns(table=table):
        if idcolumn1 in row:
            idcolumn1Found = True
    for row in excel.cursor.columns(table=table):
        if idcolumnFound == True and idcolumn1Found == True:			
            file_list = []
            sqlcommand = r"""SELECT {table}.{idcolumn}, Count({table}.OBJECTID) AS [Count]
                        FROM {table}
                        WHERE TRY_CONVERT(numeric, {idcolumn1}) = {value}
                        GROUP BY {table}.{idcolumn}
                        ORDER BY CASE WHEN {idcolumn} = TRY_CONVERT(numeric, 'NULL') THEN '1'
                                      WHEN {idcolumn} = TRY_CONVERT(numeric, 'UNK') THEN '2'
                                      ELSE {idcolumn} END ASC
                        """.format(idcolumn=idcolumn,idcolumn1=idcolumn1,table=table,value=value)
            excel.cursor.execute(sqlcommand)
            row = excel.cursor.fetchall()
            totalrows = sumrows(table)
            if not row:
                excel.cell_style = 'Bad'
                c = ""
            else:
                if int(sumrows(table)) == int(row[0][1]):
                    excel.cell_style = 'Normal'
                    cell_alignment = 'False'
                    if row[0][0] == None:
                        c = "All fields are NULL."
                    elif row[0][0] == 0:
                        c = "All fields are populated with '0'."
                    else:
                        c = "All fields are populated with '" + str(row[0][0]) +"'."
                    if (row[0][0] == None) or (row[0][0] == "") or (row[0][0] == "UNK") or (row[0] == "0E-8"):
                            excel.cell_style = 'Bad'
                else:
                    excel.cell_style = 'Normal'
                    for row in row:
                        if (row[0] == None) or (row[0] == "") or (row[0] == "UNK") or (row[0] == "0E-8"):
                            excel.cell_style = 'Bad'
                        fcol = row[0]
                        if fcol == 0:
                            fcol = "0"
                        if fcol == None:
                            fcol = "NULL"
                        fnum = row[1]
                        fper = int(round(int(fnum)/int(totalrows)*100))
                        file_list.append({"fcol": fcol, "fnum": fnum, "fper": fper})
                    cell_alignment = 'True'
                    c = r""        
                    c = c.join(("{fnum} ({fper}%) populated with '{fcol}'.\n".format(fcol=fl['fcol'], fnum=fl['fnum'], fper=fl['fper']) for fl in file_list))
                    if excel.cell_style == 'Bad':

                        r = r""
                        sqlcommand = r.join(("SELECT {table}.OBJECTID, {table}.{idcolumn}".format(idcolumn=idcolumn,table=table)))
                        if (facility_id == True and (not(idcolumn == "gs_facility_id"))):
                            sqlcommand = r.join((sqlcommand, """,{table}.gs_facility_id 
                                FROM {table}
                                WHERE ({table}.{idcolumn} IS null 
                                    OR {table}.{idcolumn} = TRY_CONVERT(numeric, '') 
                                    OR {table}.{idcolumn} LIKE '%UNK%')
                                    GROUP BY {table}.OBJECTID, {table}.{idcolumn}, {table}.gs_facility_id
                                    ORDER BY {table}.OBJECTID, {table}.{idcolumn}, {table}.gs_facility_id""".format(idcolumn=idcolumn,table=table)))
                        else:
                            sqlcommand = r.join((sqlcommand, """ FROM {table}
                                WHERE ({table}.{idcolumn} IS null 
                                    OR {table}.{idcolumn} = TRY_CONVERT(numeric, '') 
                                    OR {table}.{idcolumn} LIKE '%UNK%')
                                    GROUP BY {table}.OBJECTID, {table}.{idcolumn}
                                    ORDER BY {table}.OBJECTID, {table}.{idcolumn}""".format(idcolumn=idcolumn,table=table)))
                        df = pd.read_sql_query(sqlcommand, excel.conn)
                        sheetname = excel.category + "-" + idcolumn[len("gs_"):]
                        createsheet(sheetname)
                        writedf(df, sheetname)
                        write_hyperlink(sheetname)
            break
        else:
            excel.cell_style = 'Neutral'
            if idcolumn in row:
                c = idcolumn + " needs to be added to the database."
            elif idcolumn1 in row:
                c = idcolumn1 + " needs to be added to the database."
    
    return c

def fieldsummaryeqtext(idcolumn, idcolumn1, value, table):
    import pandas as pd
    c = "Error."
    #global excel.cell_style 
    excel.cell_style = 'Normal'
    global cell_alignment
    cell_alignment = 'False'
    facility_id = False
    idcolumnFound = False
    idcolumn1Found = False
    for row in excel.cursor.columns(table=table):
        if "gs_facility_id" in row:
            facility_id = True
    for row in excel.cursor.columns(table=table):
        if idcolumn in row:
            idcolumnFound = True
    for row in excel.cursor.columns(table=table):
        if idcolumn1 in row:
            idcolumn1Found = True
    for row in excel.cursor.columns(table=table):
        if idcolumnFound == True and idcolumn1Found == True:
            file_list = []
            sqlcommand = r"""SELECT {table}.{idcolumn}, Count({table}.OBJECTID) AS [Count]
                        FROM {table}
                        WHERE {idcolumn1} = {value}
                        GROUP BY {table}.{idcolumn}
                        ORDER BY CASE WHEN {idcolumn} = TRY_CONVERT(numeric, 'NULL') THEN '1'
                                      WHEN {idcolumn} = TRY_CONVERT(numeric, 'UNK') THEN '2'
                                      ELSE {idcolumn} END ASC
                        """.format(idcolumn=idcolumn,idcolumn1=idcolumn1,table=table,value=value)
            excel.cursor.execute(sqlcommand)
            row = excel.cursor.fetchall()
            totalrows = sumrows(table)
            if not row:
                c = ""
            else:
                if int(sumrows(table)) == int(row[0][1]):
                    excel.cell_style = 'Normal'
                    cell_alignment = 'False'
                    if row[0][0] == None:
                        c = "All fields are NULL."
                    elif row[0][0] == 0:
                        c = "All fields are populated with '0'."
                    else:
                        c = "All fields are populated with '" + str(row[0][0]) +"'."
                    if (row[0][0] == None) or (row[0][0] == "") or (row[0][0] == "UNK") or (row[0] == "0E-8"):
                            excel.cell_style = 'Bad'
                else:
                    excel.cell_style = 'Normal'
                    for row in row:
                        if (row[0] == None) or (row[0] == "") or (row[0] == "UNK") or (row[0] == "0E-8"):
                            excel.cell_style = 'Bad'
                        fcol = row[0]
                        if fcol == 0:
                            fcol = "0"
                        if fcol == None:
                            fcol = "NULL"
                        fnum = row[1]
                        fper = int(round(int(fnum)/int(totalrows)*100))
                        file_list.append({"fcol": fcol, "fnum": fnum, "fper": fper})
                    cell_alignment = 'True'
                    c = r""        
                    c = c.join(("{fnum} ({fper}%) populated with '{fcol}'.\n".format(fcol=fl['fcol'], fnum=fl['fnum'], fper=fl['fper']) for fl in file_list))
                    if excel.cell_style == 'Bad':

                        r = r""
                        sqlcommand = r.join(("SELECT {table}.OBJECTID, {table}.{idcolumn}".format(idcolumn=idcolumn,table=table)))
                        if (facility_id == True and (not(idcolumn == "gs_facility_id"))):
                            sqlcommand = r.join((sqlcommand, """,{table}.gs_facility_id 
                                FROM {table}
                                WHERE ({table}.{idcolumn} IS null 
                                    OR {table}.{idcolumn} = TRY_CONVERT(numeric, '') 
                                    OR {table}.{idcolumn} LIKE '%UNK%')
                                    GROUP BY {table}.OBJECTID, {table}.{idcolumn}, {table}.gs_facility_id
                                    ORDER BY {table}.OBJECTID, {table}.{idcolumn}, {table}.gs_facility_id""".format(idcolumn=idcolumn,table=table)))
                        else:
                            sqlcommand = r.join((sqlcommand, """ FROM {table}
                                WHERE ({table}.{idcolumn} IS null 
                                    OR {table}.{idcolumn} = TRY_CONVERT(numeric, '') 
                                    OR {table}.{idcolumn} LIKE '%UNK%')
                                    GROUP BY {table}.OBJECTID, {table}.{idcolumn}
                                    ORDER BY {table}.OBJECTID, {table}.{idcolumn}""".format(idcolumn=idcolumn,table=table)))
                        df = pd.read_sql_query(sqlcommand, excel.conn)
                        sheetname = excel.category + "-" + idcolumn[len("gs_"):]
                        createsheet(sheetname)
                        writedf(df, sheetname)
                        write_hyperlink(sheetname)
            break
        else:
            excel.cell_style = 'Neutral'
            if idcolumn in row:
                c = idcolumn + " needs to be added to the database."
            elif idcolumn1 in row:
                c = idcolumn1 + " needs to be added to the database."
    
    return c

def missingfield(idcolumn, table):
    c = "Error."
    #global excel.cell_style 
    excel.cell_style = 'Normal'
    global cell_alignment
    cell_alignment = 'False'
    for row in excel.cursor.columns(table=table):
        if idcolumn in row:
            file_list = []

            sqlcommand = r"""SELECT {table}.{idcolumn}, Count({table}.OBJECTID) AS [Count]
                        FROM {table}
                        WHERE ({table}.{idcolumn} IS null OR {table}.{idcolumn} = '')
                        GROUP BY {table}.{idcolumn}
                        ORDER BY {table}.{idcolumn}, Count({table}.OBJECTID) DESC
                        """.format(idcolumn=idcolumn,table=table)
            excel.cursor.execute(sqlcommand)
            row = excel.cursor.fetchall()
            totalrows = sumrows(table)
            
            try:
                if int(sumrows(table)) == int(row[0][1]):
                    excel.cell_style = 'Normal'
                    cell_alignment = 'False'
                    if row[0][0] == None:
                        c = "All fields are NULL."
                    elif row[0][0] == 0:
                        c = "All fields are populated with '0'."
                    else:
                        c = "All fields are populated with '" + str(row[0][0]) +"'."
                    if (row[0][0] == None) or (row[0][0] == "") or (row[0][0] == "UNK"):
                                excel.cell_style = 'Bad'
                else:
                    for row in row:
                        if (row[0] == None) or (row[0] == "") or (row[0] == "UNK"):
                            excel.cell_style = 'Bad'
                        fcol = row[0]
                        if fcol == 0:
                            fcol = "0"
                        if fcol == None:
                            fcol = "NULL"
                        fnum = row[1]
                        fper = int(round(int(fnum)/int(totalrows)*100))
                        file_list.append({"fcol": fcol, "fnum": fnum, "fper": fper})
                    cell_alignment = 'True'
                    c = r""        
                    c = c.join(("{fnum} ({fper}%) populated with '{fcol}'.\n".format(fcol=fl['fcol'], fnum=fl['fnum'], fper=fl['fper']) for fl in file_list))
                break
            except IndexError:
                break
        else:
            excel.cell_style = 'Neutral'
            c = "This field needs to be added to the database."
    
    return c

def nullabc(idcolumn, idcolumn1, idcolumn2, idcolumn3, table):
    import pandas as pd
    #global excel.cell_style 
    excel.cell_style = 'Normal'
    global cell_alignment
    cell_alignment = 'False'
    facility_id = False
    for row in excel.cursor.columns(table=table):
        if "gs_facility_id" in row:
            facility_id = True
    #for row in excel.cursor.columns(table=table):
        #if "gs_equipment_location" in row:
            #equipment_loc = True
    totalrows = sumrows(table)
    sqlcommand = r"""SELECT {table}.{idcolumn1}, {table}.{idcolumn}
                    FROM {table}
                    WHERE {table}.{idcolumn} LIKE '%a%'
                    AND ({table}.{idcolumn1} IS null OR {table}.{idcolumn1} = TRY_CONVERT(varchar,'') OR {table}.{idcolumn1} LIKE TRY_CONVERT(varchar,'%fake%') OR {table}.{idcolumn1} LIKE TRY_CONVERT(varchar,'%unk%'))
                    ORDER BY {table}.{idcolumn};
                """.format(idcolumn=idcolumn,idcolumn1=idcolumn1,table=table)

    excel.cursor.execute(sqlcommand)
    row = excel.cursor.fetchall()
    anum = len(row)
    aper = round(anum/int(totalrows)*100)
    c = str(anum) + " (" + str(aper) + "%) 'A', "

    sqlcommand = r"""SELECT {table}.{idcolumn2}, {table}.{idcolumn}
                    FROM {table}
                    WHERE {table}.{idcolumn} LIKE '%b%'
                    AND ({table}.{idcolumn2} IS null OR {table}.{idcolumn2} = TRY_CONVERT(varchar,'') OR {table}.{idcolumn2} LIKE TRY_CONVERT(varchar,'%fake%') OR {table}.{idcolumn2} LIKE TRY_CONVERT(varchar,'%unk%'))
                    ORDER BY {table}.{idcolumn};
                """.format(idcolumn=idcolumn,idcolumn2=idcolumn2,table=table)

    
    excel.cursor.execute(sqlcommand)
    row = excel.cursor.fetchall()
    bnum = len(row)
    bper = round(bnum/int(totalrows)*100)
    c = c + str(bnum) + " (" + str(bper) + "%) 'B', "
    
    sqlcommand = r"""SELECT {table}.{idcolumn3}, {table}.{idcolumn}
                    FROM {table}
                    WHERE {table}.{idcolumn} LIKE '%c%'
                    AND ({table}.{idcolumn3} IS null OR {table}.{idcolumn3} = TRY_CONVERT(varchar,'') OR {table}.{idcolumn3} LIKE TRY_CONVERT(varchar,'%fake%') OR {table}.{idcolumn3} LIKE TRY_CONVERT(varchar,'%unk%'))
                    ORDER BY {table}.{idcolumn};
                """.format(idcolumn=idcolumn,idcolumn3=idcolumn3,table=table)

    
    excel.cursor.execute(sqlcommand)
    row = excel.cursor.fetchall()
    cnum = len(row)
    cper = round(cnum/int(totalrows)*100)
    c = c + str(cnum) + " (" + str(cper) + "%) 'C' are not populated. "
    
    sqlcommand = r"""SELECT {table}.{idcolumn3}, {table}.{idcolumn}
                FROM {table}
                WHERE ({table}.{idcolumn} LIKE '%a%' OR {table}.{idcolumn} LIKE '%b%' OR {table}.{idcolumn} LIKE '%c%')
                AND ({table}.{idcolumn1} IS null OR {table}.{idcolumn1} = TRY_CONVERT(varchar,'') OR {table}.{idcolumn1} LIKE TRY_CONVERT(varchar,'%fake%') OR {table}.{idcolumn1} LIKE TRY_CONVERT(varchar,'%unk%'))
                AND ({table}.{idcolumn2} IS null OR {table}.{idcolumn2} = TRY_CONVERT(varchar,'') OR {table}.{idcolumn2} LIKE TRY_CONVERT(varchar,'%fake%') OR {table}.{idcolumn2} LIKE TRY_CONVERT(varchar,'%unk%'))
                AND ({table}.{idcolumn3} IS null OR {table}.{idcolumn3} = TRY_CONVERT(varchar,'') OR {table}.{idcolumn3} LIKE TRY_CONVERT(varchar,'%fake%') OR {table}.{idcolumn3} LIKE TRY_CONVERT(varchar,'%unk%'))
                ORDER BY {table}.{idcolumn};
            """.format(idcolumn=idcolumn,idcolumn1=idcolumn1,idcolumn2=idcolumn2,idcolumn3=idcolumn3,table=table)

    excel.cursor.execute(sqlcommand)
    row = excel.cursor.fetchall()
    tnum = len(row)
    tper = round(tnum/int(totalrows)*100)
    c = c + str(tnum) + " (" + str(tper) + "%) total."
    
    if anum + bnum + cnum == 0:
        c = "All are populated."
    else:
        excel.cell_style = 'Bad'
        r = r""
        sqlcommand = r.join(("SELECT {table}.OBJECTID,{table}.{idcolumn},{table}.{idcolumn1},{table}.{idcolumn2},{table}.{idcolumn3}".format(idcolumn=idcolumn,idcolumn1=idcolumn1,idcolumn2=idcolumn2,idcolumn3=idcolumn3,table=table)))
        if (facility_id == True):
            sqlcommand = r.join((sqlcommand, ",{table}.gs_equipment_location".format(table=table)))
        if (facility_id == True):
            sqlcommand = r.join((sqlcommand, ",{table}.gs_facility_id".format(table=table)))
        sqlcommand = r.join((sqlcommand, """ FROM {table}
                WHERE ({table}.{idcolumn} LIKE '%a%' OR {table}.{idcolumn} LIKE '%b%' OR {table}.{idcolumn} LIKE '%c%')
                AND ({table}.{idcolumn1} IS null OR {table}.{idcolumn1} = TRY_CONVERT(varchar,'') OR {table}.{idcolumn1} LIKE TRY_CONVERT(varchar,'%fake%') OR {table}.{idcolumn1} LIKE TRY_CONVERT(varchar,'%unk%'))
                AND ({table}.{idcolumn2} IS null OR {table}.{idcolumn2} = TRY_CONVERT(varchar,'') OR {table}.{idcolumn2} LIKE TRY_CONVERT(varchar,'%fake%') OR {table}.{idcolumn2} LIKE TRY_CONVERT(varchar,'%unk%'))
                AND ({table}.{idcolumn3} IS null OR {table}.{idcolumn3} = TRY_CONVERT(varchar,'') OR {table}.{idcolumn3} LIKE TRY_CONVERT(varchar,'%fake%') OR {table}.{idcolumn3} LIKE TRY_CONVERT(varchar,'%unk%'))
                ORDER BY {table}.{idcolumn};
                """.format(idcolumn=idcolumn,idcolumn1=idcolumn1,idcolumn2=idcolumn2,idcolumn3=idcolumn3,table=table)))
        df = pd.read_sql_query(sqlcommand, excel.conn)
        sheetname = excel.category + "-" + idcolumn1[len("gs_"):] + ",b,c"
        createsheet(sheetname)
        writedf(df, sheetname)
        write_hyperlink(sheetname)

    return c