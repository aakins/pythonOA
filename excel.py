def excel(xlanalysisfile):
    import openpyxl
    import pyodbc
    import os
    import pandas as pd
    import upload
    import general
    import analysis
    import analysis_special
    from datetime import date
    from pandas import ExcelWriter
    from pandas import ExcelFile
    from openpyxl.styles import Alignment
    from openpyxl.utils.dataframe import dataframe_to_rows
    global conn
    global category
    global wb
    global ws
    global cell
    global cell_alignment
    global cell_style
    global idcolumn
    global idcolumn1
    global idcolumn2
    global idcolumn3
    conn = pyodbc.connect("Driver={driver};Server=.\SQLEXPRESS;Database={database};Trusted_Connection=yes".format(driver = "{SQL Server}",database = "gs" + general.number), autocommit = True)
    conn.timeout = 60
    #conn.setencoding('utf-8')  # (Python 3.x syntax)
    #conn.setdecoding(pyodbc.SQL_CHAR, encoding='utf-8')
    #conn.setdecoding(pyodbc.SQL_WCHAR, encoding='utf-8')
    #conn.setencoding(encoding='utf-8')
    global cursor
    cursor = conn.cursor()

    cell_style = 'Normal'
    cell_alignment = 'False'
    cell = 2

    wb = openpyxl.load_workbook(xlanalysisfile, keep_vba=True)
    ws = wb["Overview"]

    c = "OA Data Analysis Summary \n" + general.number
    ws['A1'] = c
    c = "Review of Snapshot from " + str(date.today())
    ws['C1'] = c

    category = "SUB"
    #copysubstation = 0
    table = "gs_electric_station"
    idcolumn = "gs_name"
    c = "Substations (" + str(analysis.sumrows (table)) + ")"
    ws['A'+ str(cell)] = c
    cell += 1
    c = analysis.uniqueid(idcolumn,table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)
    cell += 1
    idcolumn1 = "gs_facility_id"
    c = analysis.duplicateid(idcolumn, idcolumn1, table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)
    idcolumn = "gs_phase"
    cell += 1
    c = analysis.fieldsummary(idcolumn,table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)
    idcolumn = "gs_rated_voltage"
    cell += 1
    c = analysis.fieldsummary(idcolumn,table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)
    idcolumn = "gs_connection_code"
    cell += 1
    c = analysis.fieldsummary(idcolumn,table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)
    idcolumn = "gs_positive_r"
    cell += 1
    c = analysis.fieldsummary(idcolumn,table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)
    idcolumn = "gs_positive_x"
    cell += 1
    c = analysis.fieldsummary(idcolumn,table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)
    idcolumn = "gs_zero_r"
    cell += 1
    c = analysis.fieldsummary(idcolumn,table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)
    idcolumn = "gs_zero_x"
    cell += 1
    c = analysis.fieldsummary(idcolumn,table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)
    idcolumn = "gs_bus_voltage"
    cell += 1
    c = analysis.fieldsummary(idcolumn,table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)

    category = "CAP"
    table = "gs_capacitor_bank"
    idcolumn = "gs_equipment_location"
    c = "Capacitors (" + str(analysis.sumrows (table)) + ")"
    cell += 2
    ws['A'+ str(cell)] = c
    cell += 1
    c = analysis.uniqueid(idcolumn,table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)
    idcolumn = "gs_equipment_location"
    idcolumn1 = "gs_facility_id"
    cell += 1
    c = analysis.duplicateid(idcolumn, idcolumn1, table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)
    idcolumn = "gs_phase"
    cell += 1
    c = analysis.fieldsummary(idcolumn,table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)
    idcolumn = "gs_unit_size_kvar"
    cell += 1
    c = analysis.fieldsummary(idcolumn,table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)
    idcolumn = "gs_voltage_rating"
    cell += 1
    c = analysis.fieldsummary(idcolumn,table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)
    idcolumn = "gs_status_code"
    cell += 1
    c = analysis.fieldsummary(idcolumn,table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)
    idcolumn = "gs_type_code"
    cell += 1
    c = analysis.fieldsummary(idcolumn,table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)
    idcolumn = "gs_connection"
    cell += 1
    c = analysis.fieldsummary(idcolumn,table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)
    idcolumn = "gs_control_element"
    idcolumn1 = "gs_type_code"
    value = 0
    cell += 1
    c = analysis.fieldsummarygt(idcolumn, idcolumn1, value, table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)
    idcolumn = "gs_gang_controlled"
    cell += 1
    c = analysis.fieldsummary(idcolumn,table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)
    idcolumn = "gs_control_phase"
    idcolumn1 = "gs_gang_controlled"
    value = 1
    cell += 1
    c = analysis.fieldsummaryeq(idcolumn, idcolumn1, value, table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)
    idcolumn = "gs_on_set"
    idcolumn1 = "gs_type_code"
    value = 0
    cell += 1
    c = analysis.fieldsummarygt(idcolumn, idcolumn1, value, table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)
    idcolumn = "gs_off_set"
    idcolumn1 = "gs_type_code"
    value = 0
    cell += 1
    c = analysis.fieldsummarygt(idcolumn, idcolumn1, value, table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)
    idcolumn = "gs_on_set_winter"
    idcolumn1 = "gs_type_code"
    value = 2
    cell += 1
    c = analysis.fieldsummaryeq(idcolumn, idcolumn1, value, table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)
    idcolumn = "gs_off_set_winter"
    idcolumn1 = "gs_type_code"
    value = 2
    cell += 1
    c = analysis.fieldsummaryeq(idcolumn, idcolumn1, value, table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)
    idcolumn = "gs_on_month"
    idcolumn1 = "gs_type_code"
    value = 4
    cell += 1
    c = analysis.fieldsummarygt(idcolumn, idcolumn1, value, table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)
    idcolumn = "gs_on_day"
    idcolumn1 = "gs_type_code"
    value = 4
    cell += 1
    c = analysis.fieldsummarygt(idcolumn, idcolumn1, value, table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)
    idcolumn = "gs_cust_volts_override"
    idcolumn1 = "gs_type_code"
    value = 5
    cell += 1
    c = analysis.fieldsummarygt(idcolumn, idcolumn1, value, table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)
    idcolumn = "gs_min_volts_override"
    cell += 1
    c = analysis.fieldsummary(idcolumn,table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)
    idcolumn = "gs_max_volts_override"
    cell += 1
    c = analysis.fieldsummary(idcolumn,table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)

    category = "OCD"
    table = "gs_overcurrent_device"
    idcolumn = "gs_equipment_location"
    c = "Overcurrent Devices (" + str(analysis.sumrows (table)) + ")"
    cell += 2
    ws['A'+ str(cell)] = c
    cell += 1
    c = analysis.uniqueid(idcolumn,table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)
    idcolumn = "gs_equipment_location"
    idcolumn1 = "gs_phase"
    cell += 1
    c = analysis.duplicateid(idcolumn, idcolumn1, table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)
    idcolumn = "gs_phase"
    cell += 1
    c = analysis.fieldsummary(idcolumn,table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)
    idcolumn = "gs_phase"
    idcolumn1 = "gs_device_desc_a"
    idcolumn2 = "gs_device_desc_b"
    idcolumn3 = "gs_device_desc_c"
    cell += 1
    c = analysis.nullabc(idcolumn, idcolumn1, idcolumn2, idcolumn3, table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)
    idcolumn = "gs_overcurrent_device_subtype"
    cell += 1
    c = analysis.fieldsummary(idcolumn,table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)
    idcolumn = "gs_switch_description"
    idcolumn1 = "gs_overcurrent_device_subtype"
    value = 1
    cell += 1
    c = analysis.fieldsummaryeq(idcolumn,idcolumn1, value, table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)
    idcolumn = "gs_equipment_location"
    cell += 1
    c = analysis_special.is_feeder_bay(idcolumn,table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)

    category = "SWI"
    table = "gs_switch"
    idcolumn = "gs_equipment_location"
    c = "Switches (" + str(analysis.sumrows (table)) + ")"
    cell += 3
    ws['A'+ str(cell)] = c
    cell += 1
    c = analysis.uniqueid(idcolumn,table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)
    idcolumn = "gs_equipment_location"
    idcolumn1 = "gs_facility_id"
    cell += 1
    c = analysis.duplicateid(idcolumn, idcolumn1, table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)
    idcolumn = "gs_phase"
    cell += 1
    c = analysis.fieldsummary(idcolumn,table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)
    idcolumn = "gs_switch_status"
    cell += 1
    c = analysis.fieldsummary(idcolumn,table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)
    idcolumn = "gs_switch_description"
    cell += 1
    c = analysis.fieldsummary(idcolumn,table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)

    category = "XFMR"
    table = "gs_transformer"
    idcolumn = "gs_equipment_location"
    c = "Transformers (" + str(analysis.sumrows (table)) + ")"
    cell += 2
    ws['A'+ str(cell)] = c
    cell += 1
    c = analysis.uniqueid(idcolumn,table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)
    idcolumn = "gs_equipment_location"
    idcolumn1 = "gs_bank_id"
    cell += 1
    c = analysis_special.duplicate_xfmr(idcolumn, idcolumn1, table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)
    idcolumn = "gs_phase"
    cell += 1
    c = analysis.fieldsummary(idcolumn,table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)
    idcolumn = "gs_xfmr_conductor_description"
    cell += 1
    c = analysis.fieldsummary(idcolumn,table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)
    idcolumn = "gs_phase"
    idcolumn1 = "gs_tran_kva_a"
    idcolumn2 = "gs_tran_kva_b"
    idcolumn3 = "gs_tran_kva_c"
    cell += 1
    c = analysis.nullabc(idcolumn, idcolumn1, idcolumn2, idcolumn3, table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)
    idcolumn = "gs_winding_connection"
    cell += 1
    c = analysis.fieldsummary(idcolumn,table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)
    idcolumn = "gs_rated_input_voltage"
    cell += 1
    c = analysis.fieldsummary(idcolumn,table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)
    idcolumn = "gs_rated_input_voltage"
    idcolumn1 = "gs_rated_output_voltage"
    analysis_special.xfmr_voltage(idcolumn,idcolumn1,table)
    idcolumn = "gs_rated_output_voltage"
    cell += 1
    c = analysis.fieldsummary(idcolumn,table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)
    #idcolumn = "gs_equipment_location"
    #idcolumn1 = "gs_bank_id"
    #c = analysis.duplicateid(idcolumn, idcolumn1, table) + " They should have gs_bank_ids added if they are banked."
    #ws['C45'] = c
    #ws['C45'].style = cell_style
    #ws['C45'].alignment = Alignment(wrap_text=cell_alignment)
    idcolumn = "gs_is_substation_transformer"
    cell += 2
    c = analysis.fieldsummary(idcolumn,table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)
    idcolumn = "gs_positive_r"
    idcolumn1 = "gs_is_substation_transformer"
    value = "\'true\'"
    cell += 1
    c = analysis.fieldsummaryeqtext(idcolumn, idcolumn1, value, table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)
    idcolumn = "gs_positive_x"
    idcolumn1 = "gs_is_substation_transformer"
    value = "\'true\'"
    cell += 1
    c = analysis.fieldsummaryeqtext(idcolumn, idcolumn1, value, table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)
    idcolumn = "gs_zero_r"
    idcolumn1 = "gs_is_substation_transformer"
    value = "\'true\'"
    cell += 1
    c = analysis.fieldsummaryeqtext(idcolumn, idcolumn1, value, table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)
    idcolumn = "gs_zero_x"
    idcolumn1 = "gs_is_substation_transformer"
    value = "\'true\'"
    cell += 1
    c = analysis.fieldsummaryeqtext(idcolumn, idcolumn1, value, table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)
    idcolumn = "gs_impedance"
    idcolumn1 = "gs_is_substation_transformer"
    value = "\'true\'"
    cell += 1
    c = analysis.fieldsummaryeqtext(idcolumn, idcolumn1, value, table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)
    idcolumn = "gs_impedance_angle"
    idcolumn1 = "gs_is_substation_transformer"
    value = "\'true\'"
    cell += 1
    c = analysis.fieldsummaryeqtext(idcolumn, idcolumn1, value, table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)
    idcolumn = "gs_is_center_tap"
    idcolumn1 = "gs_is_substation_transformer"
    value = "\'true\'"
    cell += 1
    c = analysis.fieldsummaryeqtext(idcolumn, idcolumn1, value, table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)

    category = "VREG"
    table = "gs_voltage_regulator"
    idcolumn = "gs_equipment_location"
    c = "Voltage Regulators (" + str(analysis.sumrows (table)) + ")"
    cell += 2
    ws['A'+ str(cell)] = c
    cell += 1
    c = analysis.uniqueid(idcolumn,table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)
    idcolumn = "gs_equipment_location"
    idcolumn1 = "gs_facility_id"
    cell += 1
    c = analysis.duplicateid(idcolumn, idcolumn1, table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)
    idcolumn = "gs_phase"
    cell += 1
    c = analysis.fieldsummary(idcolumn,table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)
    idcolumn = "gs_phase"
    idcolumn1 = "gs_regulator_a"
    idcolumn2 = "gs_regulator_b"
    idcolumn3 = "gs_regulator_c"
    cell += 1
    c = analysis.nullabc(idcolumn, idcolumn1, idcolumn2, idcolumn3, table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)
    idcolumn = "gs_winding_connection"
    cell += 1
    c = analysis.fieldsummary(idcolumn,table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)
    idcolumn = "gs_nominal_voltage"
    cell += 1
    c = analysis.fieldsummary(idcolumn,table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)
    idcolumn = "gs_base_volts"
    cell += 1
    c = analysis.fieldsummary(idcolumn,table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)
    idcolumn = "gs_bandwidth"
    cell += 1
    c = analysis.fieldsummary(idcolumn,table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)
    idcolumn = "gs_ldc_a_total"
    cell += 1
    c = analysis.fieldsummary(idcolumn,table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)
    idcolumn = "gs_ldc_r_total"
    cell += 1
    c = analysis.fieldsummary(idcolumn,table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)
    idcolumn = "gs_ldc_x_total"
    cell += 1
    c = analysis.fieldsummary(idcolumn,table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)
    idcolumn = "gs_regulator_mode"
    cell += 1
    c = analysis.fieldsummary(idcolumn,table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)
    idcolumn = "gs_step_a"
    cell += 1
    c = analysis.fieldsummary(idcolumn,table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)
    idcolumn = "gs_step_b"
    cell += 1
    c = analysis.fieldsummary(idcolumn,table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)
    idcolumn = "gs_step_c"
    cell += 1
    c = analysis.fieldsummary(idcolumn,table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)
    idcolumn = "gs_block_step"
    cell += 1
    c = analysis.fieldsummary(idcolumn,table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)
    idcolumn = "gs_regulator_type"
    cell += 1
    c = analysis.fieldsummary(idcolumn,table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)
    idcolumn = "gs_controlling_phase"
    idcolumn1 = "gs_regulator_type"
    value = 1
    cell += 1
    c = analysis.fieldsummaryeq(idcolumn, idcolumn1, value, table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)
    idcolumn = "gs_regulating_phase"
    cell += 1
    c = analysis.fieldsummary(idcolumn,table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)

    category = "SPAN"
    table = "gs_span"
    c = "Conductors (" + str(analysis.sumrows (table)) + ")"
    cell += 2
    ws['A'+ str(cell)] = c
    idcolumn = "gs_phase"
    cell += 1
    c = analysis.fieldsummary(idcolumn,table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)
    idcolumn = "gs_phase"
    idcolumn1 = "gs_conductor_a"
    idcolumn2 = "gs_conductor_b"
    idcolumn3 = "gs_conductor_c"
    cell += 1
    c = analysis.nullabc(idcolumn, idcolumn1, idcolumn2, idcolumn3, table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)
    idcolumn = "gs_conductor_n"
    cell += 1
    c = analysis_special.neutral(idcolumn,table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)
    idcolumn = "gs_subtype_cd"
    cell += 1
    c = analysis.fieldsummary(idcolumn,table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=True)
    idcolumn = "gs_construction_desc"
    cell += 1
    c = analysis.missingfield(idcolumn,table)
    ws['C'+ str(cell)] = c + " This is common in GIS data and we compensate by populating that field with RUS standards for the most common configurations."
    ws['C'+ str(cell)].style = "Normal"
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)
    ws.column_dimensions['C'].auto_size = True
    cell += 1
    analysis_special.span_assembly(table)

    category = "MTR"
    table = "gs_motor"
    c = "Motors (" + str(analysis.sumrows (table)) + ")"
    cell += 2
    ws['A'+ str(cell)] = c
    idcolumn = "gs_equipment_location"
    cell += 1
    c = analysis.uniqueid(idcolumn,table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)
    idcolumn = "gs_nema_type"
    cell += 1
    c = analysis.fieldsummary(idcolumn,table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)
    idcolumn = "gs_soft_start_tap"
    cell += 1
    c = analysis.fieldsummary(idcolumn,table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)
    idcolumn = "gs_rated_hp"
    cell += 1
    c = analysis.fieldsummary(idcolumn,table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)
    idcolumn = "gs_power_factor"
    cell += 1
    c = analysis.fieldsummary(idcolumn,table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)
    idcolumn = "gs_locked_rotor_multiplier"
    cell += 1
    c = analysis.fieldsummary(idcolumn,table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)
    idcolumn = "gs_locked_rotor_pf"
    cell += 1
    c = analysis.fieldsummary(idcolumn,table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)

    category = "GEN"
    table = "gs_generator"
    c = "Generators (" + str(analysis.sumrows (table)) + ")"
    cell += 2
    ws['A'+ str(cell)] = c
    idcolumn = "gs_subtype_cd"
    cell += 1
    c = analysis.fieldsummary(idcolumn,table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)
    idcolumn = "gs_max_kw_out"
    cell += 1
    c = analysis.fieldsummary(idcolumn,table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)
    idcolumn = "gs_rated_kva"
    cell += 1
    c = analysis.fieldsummary(idcolumn,table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)
    idcolumn = "gs_max_kvar_lagg"
    cell += 1
    c = analysis.fieldsummary(idcolumn,table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)
    idcolumn = "gs_power_factor"
    cell += 1
    c = analysis.fieldsummary(idcolumn,table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)
    idcolumn = "gs_on_off"
    cell += 1
    c = analysis.fieldsummary(idcolumn,table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)
    idcolumn = "gs_installation_date"
    cell += 1
    #c = analysis.fieldsummary(idcolumn,table)
    #ws['C'+ str(cell)] = c
    #ws['C'+ str(cell)].style = cell_style
    #ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)
    idcolumn = "gs_fault_contribution"
    cell += 1
    c = analysis.fieldsummary(idcolumn,table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)
    idcolumn = "gs_inverter_efficiency"
    cell += 1
    c = analysis.fieldsummary(idcolumn,table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)
    idcolumn = "gs_curtailing_component_id"
    cell += 1
    c = analysis.fieldsummary(idcolumn,table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)
    idcolumn = "gs_power_factor_response"
    cell += 1
    c = analysis.fieldsummary(idcolumn,table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)
    idcolumn = "gs_rpm"
    cell += 1
    c = analysis.fieldsummary(idcolumn,table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)
    idcolumn = "gs_mva_base"
    cell += 1
    c = analysis.fieldsummary(idcolumn,table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)
    idcolumn = "gs_kva_base"
    cell += 1
    c = analysis.fieldsummary(idcolumn,table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)
    idcolumn = "gs_positive_sequence_reactance"
    cell += 1
    c = analysis.fieldsummary(idcolumn,table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)
    idcolumn = "gs_saturated_sequence_reactance"
    cell += 1
    c = analysis.fieldsummary(idcolumn,table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)
    idcolumn = "gs_transient_reactance"
    cell += 1
    c = analysis.fieldsummary(idcolumn,table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)
    idcolumn = "gs_subtransient_reactance"
    cell += 1
    c = analysis.fieldsummary(idcolumn,table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)
    idcolumn = "gs_transient_time_constant"
    cell += 1
    c = analysis.fieldsummary(idcolumn,table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)
    idcolumn = "gs_subtransient_time_constant"
    cell += 1
    c = analysis.fieldsummary(idcolumn,table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)
    idcolumn = "gs_num_poles"
    cell += 1
    c = analysis.fieldsummary(idcolumn,table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)
    idcolumn = "gs_slip"
    cell += 1
    c = analysis.fieldsummary(idcolumn,table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)
    idcolumn = "gs_stator_resistance"
    cell += 1
    c = analysis.fieldsummary(idcolumn,table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)
    idcolumn = "gs_stator_reactance"
    cell += 1
    c = analysis.fieldsummary(idcolumn,table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)
    idcolumn = "gs_rotor_resistance"
    cell += 1
    c = analysis.fieldsummary(idcolumn,table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)
    idcolumn = "gs_rotor_reactance"
    cell += 1
    c = analysis.fieldsummary(idcolumn,table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)
    idcolumn = "gs_magnetizing_reactance"
    cell += 1
    c = analysis.fieldsummary(idcolumn,table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)
    idcolumn = "gs_crowbar_resistance"
    cell += 1
    c = analysis.fieldsummary(idcolumn,table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)
    idcolumn = "gs_tilt"
    cell += 1
    c = analysis.fieldsummary(idcolumn,table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)
    idcolumn = "gs_azimuth"
    cell += 1
    c = analysis.fieldsummary(idcolumn,table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)
    idcolumn = "gs_cpr_site_id"
    cell += 1
    c = analysis.fieldsummary(idcolumn,table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)
    idcolumn = "gs_charge"
    cell += 1
    c = analysis.fieldsummary(idcolumn,table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)
    idcolumn = "gs_kw"
    cell += 1
    c = analysis.fieldsummary(idcolumn,table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)
    idcolumn = "gs_cell_voltage"
    cell += 1
    c = analysis.fieldsummary(idcolumn,table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)
    idcolumn = "gs_cell_resistance"
    cell += 1
    c = analysis.fieldsummary(idcolumn,table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)
    idcolumn = "gs_cell_count"
    cell += 1
    c = analysis.fieldsummary(idcolumn,table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)
    idcolumn = "gs_blade_diameter"
    cell += 1
    c = analysis.fieldsummary(idcolumn,table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)
    idcolumn = "gs_performance_coefficient"
    cell += 1
    c = analysis.fieldsummary(idcolumn,table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)
    idcolumn = "gs_generator_efficiency"
    cell += 1
    c = analysis.fieldsummary(idcolumn,table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)
    idcolumn = "gs_gear_box_efficiency"
    cell += 1
    c = analysis.fieldsummary(idcolumn,table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)

    category = "SERV"
    table = "gs_service_point"
    c = "Service Points (" + str(analysis.sumrows (table)) + ")"
    cell += 2
    ws['A'+ str(cell)] = c
    cell += 1
    idcolumn = "gs_service_map_location"
    c = analysis.uniqueid(idcolumn,table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)
    idcolumn1 = "gs_service_number"
    cell += 1
    c = analysis.duplicateid(idcolumn, idcolumn1, table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)
    idcolumn = "gs_phase"
    cell += 1
    c = analysis.fieldsummary(idcolumn,table)
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)
    cell += 1
    c = analysis_special.duplicate_system()
    ws['C'+ str(cell)] = c
    ws['C'+ str(cell)].style = cell_style
    ws['C'+ str(cell)].alignment = Alignment(wrap_text=cell_alignment)

    print("Done.")
    conn.close()
    wb.save(xlanalysisfile)
    wb.close()
    return